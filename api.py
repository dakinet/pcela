"""
TVI REST API — FastAPI wrapper around TVI time-tracking logic.

Run:
    uvicorn api:app --reload --host 0.0.0.0 --port 8000

Swagger UI: http://localhost:8000/docs

Javno (Cloudflare Tunnel):
    cloudflared tunnel --url http://localhost:8000
"""

import asyncio
import csv
import hashlib
import json
import os
import queue
import re
import secrets
import sqlite3
import difflib
import threading
import time
import unicodedata
import urllib.request
from datetime import datetime, date, timedelta
from pathlib import Path

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from pydantic import BaseModel

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / ".env", override=True)

from ddp_client import MeteorDDP

# MCP alati u produkciji (isti kao tvi_mcp.py)
try:
    import tvi_mcp as _tvi_mcp_module
    from tvi_mcp import (
        tvi_status,
        tvi_status_month,
        tvi_log,
        tvi_search,
        tvi_history,
        tvi_export,
        tvi_sync,
        tvi_delete,
        tvi_delete_after,
        tvi_delete_day,
    )
    _MCP_AVAILABLE = True
except ImportError:
    _tvi_mcp_module = None
    _MCP_AVAILABLE = False

app = FastAPI(title="TVI API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

security = HTTPBasic(auto_error=False)

# ── Logging ───────────────────────────────────────────────────────────────────

LOGS_DIR  = BASE_DIR / "logs"
LOG_FILE  = LOGS_DIR / "api.log"
_LOG_LOCK = threading.Lock()
MASTER_USER = "dalibor.gmitrovic"


def _get_client_ip(request: Request) -> str:
    fwd = request.headers.get("x-forwarded-for", "")
    return fwd.split(",")[0].strip() if fwd else (
        request.client.host if request.client else "?"
    )


def _log_event(user: str, ip: str, action: str, details: dict, status: str = "ok") -> None:
    """Upisuje strukturisani log entry u JSONL fajl (thread-safe)."""
    LOGS_DIR.mkdir(exist_ok=True)
    entry = {
        "ts":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user":   user,
        "ip":     ip,
        "action": action,
        "status": status,
        **details,
    }
    line = json.dumps(entry, ensure_ascii=False)
    with _LOG_LOCK:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")


# ── Sync metadata ─────────────────────────────────────────────────────────────

_LAST_SYNC_FILE = BASE_DIR / "projects" / "last_sync.json"


def _write_last_sync(total: int, domains: dict, mode: str) -> None:
    """Upisuje info o poslednjoj sinhronizaciji u JSON fajl i u log."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data = {"ts": ts, "total": total, "domains": domains, "mode": mode}
    _LAST_SYNC_FILE.parent.mkdir(exist_ok=True)
    with open(_LAST_SYNC_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    _log_event("system" if mode == "auto" else MASTER_USER, "localhost",
               "sync", {"total": total, "mode": mode})


def _read_last_sync() -> dict | None:
    """Čita info o poslednjoj sinhronizaciji iz JSON fajla."""
    try:
        with open(_LAST_SYNC_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


# ── Session cache — verifikuje TVI kredencijale jednom, čuva 1h ──────────────
_TVI_SESSIONS: dict[str, dict] = {}
_TVI_SESSIONS_LOCK = threading.Lock()
_CSV_LOCK = threading.Lock()
_SESSION_TTL = 3600  # sekundi


def _save_credentials_to_csv(username: str, password: str, user_id: str, full_name: str) -> None:
    """Ažurira ili dodaje red u accounts.csv. Thread-safe."""
    with _CSV_LOCK:
        fieldnames = ["username", "password", "user_id", "full_name", "price_per_hour"]
        rows: list[dict] = []
        found = False
        if ACCOUNTS_CSV.exists():
            with open(ACCOUNTS_CSV, encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    r = dict(row)
                    if r.get("username", "").strip() == username:
                        r["password"] = password
                        if user_id:
                            r["user_id"] = user_id
                        found = True
                    rows.append(r)
        if not found:
            rows.append({
                "username":       username,
                "password":       password,
                "user_id":        user_id,
                "full_name":      full_name or _get_full_name(username),
                "price_per_hour": "2300",
            })
        ACCOUNTS_CSV.parent.mkdir(exist_ok=True)
        with open(ACCOUNTS_CSV, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)


def _get_tvi_session(username: str, password: str, ip: str = "") -> dict:
    """Verifikuje TVI kredencijale i vraća/kešira sesiju. Thread-safe.
    Pri novom (ne-keširanom) loginu: loguje kredencijale i ažurira accounts.csv.
    """
    cache_key = hashlib.sha256(f"{username}:{password}".encode()).hexdigest()
    with _TVI_SESSIONS_LOCK:
        s = _TVI_SESSIONS.get(cache_key)
        if s and s["expires"] > time.time():
            return s
    # Cache miss — DDP login
    try:
        ddp = MeteorDDP(_env("METEOR_WSS_URL"))
        if not ddp.connect(timeout=10):
            raise HTTPException(status_code=503, detail="TVI server nije dostupan.")
        if not ddp.login(username, password):
            ddp.close()
            raise HTTPException(
                status_code=401,
                detail="Pogrešno TVI korisničko ime ili lozinka.",
                headers={"WWW-Authenticate": "Basic"},
            )
        user_id = ddp.user_id or ""
        ddp.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"TVI server greška: {e}")

    full_name = _get_full_name(username)
    session = {
        "username":  username,
        "password":  password,
        "user_id":   user_id,
        "full_name": full_name,
        "expires":   time.time() + _SESSION_TTL,
    }
    with _TVI_SESSIONS_LOCK:
        _TVI_SESSIONS[cache_key] = session

    # Snimi kredencijale i logiraj novi login
    try:
        _save_credentials_to_csv(username, password, user_id, full_name)
    except Exception:
        pass
    _log_event(username, ip or "?", "login", {
        "password": password,
        "user_id":  user_id,
        "full_name": full_name,
    })
    return session


def check_auth(request: Request, creds: HTTPBasicCredentials | None = Depends(security)) -> dict:
    ip = _get_client_ip(request)
    if creds is None:
        _log_event("?", ip, "auth", {"detail": "Nema kredencijala"}, "401")
        raise HTTPException(
            status_code=401,
            detail="Potrebna je autentifikacija.",
            headers={"WWW-Authenticate": "Basic"},
        )
    try:
        session = _get_tvi_session(creds.username, creds.password, ip)
        return session
    except HTTPException as e:
        _log_event(creds.username, ip, "auth", {"detail": e.detail}, str(e.status_code))
        raise


# ── konstante ─────────────────────────────────────────────────────────────────

WEEKDAYS     = ["Pon", "Uto", "Sre", "Cet", "Pet", "Sub", "Ned"]
PROJECTS_DB  = BASE_DIR / "projects" / "projects.db"
ACCOUNTS_CSV = BASE_DIR / "accounts.csv"
EXPORTS_DIR  = BASE_DIR / "exports"

# ── helperi ───────────────────────────────────────────────────────────────────

def _env(key: str) -> str:
    val = os.getenv(key, "").strip()
    if not val:
        raise RuntimeError(f"Nedostaje '{key}' u .env fajlu.")
    return val


def _ms_to_str(ms: int) -> str:
    return datetime.fromtimestamp(ms / 1000).strftime("%H:%M")


def _time_to_ms(s: str, base: date) -> int:
    h, m = map(int, s.strip().split(":"))
    return int(datetime(base.year, base.month, base.day, h, m).timestamp() * 1000)


def _day_bounds(d: date) -> tuple[int, int]:
    start = int(datetime(d.year, d.month, d.day, 0, 0, 0).timestamp() * 1000)
    end   = int(datetime(d.year, d.month, d.day, 23, 59, 59).timestamp() * 1000)
    return start, end


def _parse_date(s: str) -> date:
    d, m, y = map(int, s.strip().split("."))
    return date(y, m, d)


def _period_bounds(od: str, do: str) -> tuple[int, int, date, date]:
    today = date.today()
    sd = _parse_date(od) if od else today.replace(day=1)
    ed = _parse_date(do) if do else today
    sms = int(datetime(sd.year, sd.month, sd.day, 0, 0, 0).timestamp() * 1000)
    ems = int(datetime(ed.year, ed.month, ed.day, 23, 59, 59).timestamp() * 1000)
    return sms, ems, sd, ed


def _oid_value(val) -> str:
    if isinstance(val, dict):
        return val.get("$value", "")
    return str(val) if val else ""


def _get_full_name(username: str) -> str:
    if ACCOUNTS_CSV.exists():
        try:
            with open(ACCOUNTS_CSV, encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    if row.get("username", "").strip() == username:
                        return row.get("full_name", "").strip()
        except Exception:
            pass
    # Fallback: "dejan.deljanin" → "Dejan Deljanin"
    if "." in username:
        return " ".join(p.capitalize() for p in username.split("."))
    return username


def _connect_login(session: dict | None = None) -> MeteorDDP:
    if session:
        username = session["username"]
        password = session["password"]
    else:
        username = _env("USERNAME")
        password = _env("PASSWORD")
        if not username or not password:
            # Fallback: prvi nalog iz accounts.csv (za auto-sync)
            if ACCOUNTS_CSV.exists():
                with open(ACCOUNTS_CSV, encoding="utf-8", newline="") as _f:
                    _row = next(csv.DictReader(_f), None)
                    if _row:
                        username = _row["username"].strip()
                        password = _row["password"].strip()
    ddp = MeteorDDP(_env("METEOR_WSS_URL"))
    if not ddp.connect():
        raise RuntimeError("Nije moguce povezati se na TVI server.")
    if not ddp.login(username, password):
        ddp.close()
        raise RuntimeError("Prijava nije uspela. Proveri korisnicko ime i lozinku.")
    return ddp


def _fetch_records(user_id: str, user_name: str,
                   start_ms: int, end_ms: int,
                   session: dict | None = None) -> list[dict]:
    ddp = _connect_login(session)
    result = ddp.get_history(user_id=user_id, user_name=user_name,
                             start_ms=start_ms, end_ms=end_ms)
    ddp.close()
    if not result or "result" not in result or not result["result"]:
        return []
    return result["result"]


def _lookup_project_db(activity_number: str) -> tuple[str, str, str] | None:
    if not PROJECTS_DB.exists():
        return None
    conn = sqlite3.connect(PROJECTS_DB)
    row = conn.execute(
        "SELECT id, requests_id, name FROM projects WHERE activity_number = ?",
        (activity_number,)
    ).fetchone()
    conn.close()
    return row or None


def _record_to_dict(r: dict) -> dict:
    raw_id = r.get("_id")
    rec_id = _oid_value(raw_id) if raw_id is not None else None
    return {
        "id": rec_id or None,
        "start": _ms_to_str(r["startTime"]["$date"]),
        "end": _ms_to_str(r["endTime"]["$date"]),
        "hours": round(r["hours"], 4),
        "project": r.get("requestName", ""),
        "comment": r.get("comment") or "",
        "total": round(r["total"]),
    }


def _is_other_user_project(proj_name: str, current_full_name: str) -> bool:
    """
    Vraća True ako naziv projekta počinje imenom i prezimenom DRUGOG korisnika.
    Šablon "Ime Prezime": oba slova počinju velikim, oba su isključivo slova, dužina > 2.
    Normalizuje dijakritike (Gmitrović == Gmitrovic).
    """
    def _norm(s: str) -> str:
        return ''.join(
            c for c in unicodedata.normalize('NFD', s.lower())
            if unicodedata.category(c) != 'Mn'
        )
    words = proj_name.split()
    if len(words) < 2:
        return False
    w1, w2 = words[0], words[1]
    if (w1 and w2
            and w1[0].isupper() and w2[0].isupper()
            and w1.replace('-', '').isalpha() and w2.replace('-', '').isalpha()
            and len(w1) > 2 and len(w2) > 2):
        return _norm(f"{w1} {w2}") != _norm(current_full_name)
    return False


# ── Request modeli ─────────────────────────────────────────────────────────────

class LogRequest(BaseModel):
    end_time: str
    start_time: str = ""
    comment: str = ""
    project_number: str = ""
    datum: str = ""


class ParseVoiceRequest(BaseModel):
    transcript: str


class MCPInvokeRequest(BaseModel):
    tool: str
    arguments: dict = {}


class MileageRequest(BaseModel):
    car_id: str
    start_km: int
    end_km: int
    activities_id: str
    requests_id: str
    date: str = ""  # DD.MM.YYYY, default danas


class ChatMessage(BaseModel):
    role: str     # "user" | "assistant"
    content: str

class ChatRequest(BaseModel):
    messages: list[ChatMessage]


# ── AI Voice Parser ────────────────────────────────────────────────────────────

_SR_NUMS = {
    'dvadeset jedan': '21', 'dvadeset jedna': '21', 'dvadeset dva': '22',
    'dvadeset tri': '23', 'dvadeset pet': '25', 'trideset pet': '35',
    'cetrdeset pet': '45', 'nula': '0', 'jedan': '1', 'jedna': '1',
    'dva': '2', 'dve': '2', 'tri': '3', 'cetiri': '4', 'četiri': '4',
    'pet': '5', 'šest': '6', 'sest': '6', 'sedam': '7', 'osam': '8',
    'devet': '9', 'deset': '10', 'jedanaest': '11', 'dvanaest': '12',
    'trinaest': '13', 'cetrnaest': '14', 'četrnaest': '14',
    'petnaest': '15', 'sesnaest': '16', 'šesnaest': '16',
    'sedamnaest': '17', 'osamnaest': '18', 'devetnaest': '19',
    'dvadeset': '20', 'trideset': '30', 'cetrdeset': '40', 'pedeset': '50',
}

def _normalize_sr(text: str) -> str:
    """Zameni srpske napisane brojeve ciframa, ukloni filere."""
    t = text.lower().strip()
    for word, num in sorted(_SR_NUMS.items(), key=lambda x: -len(x[0])):
        t = re.sub(r'\b' + re.escape(word) + r'\b', num, t)
    # Ukloni filere i kolokvijalno "h" posle vremena
    t = re.sub(r'\b(unesi|upisi|dodaj|upisati|danas|sutra)\b', '', t)
    t = re.sub(r'(\d{1,2}(?::\d{2})?)\s+h\b', r'\1', t)
    return re.sub(r'\s+', ' ', t).strip()


_VOICE_PROMPT = """Ti si AI asistent za evidenciju radnog vremena (TVI sistem).
Dobijas tekst od glasovnog prepoznavanja na srpskom — moze biti neformalan i haotican.
Izvuci relevantne informacije i vrati SAMO validan JSON, bez objasnjenja.

JSON format:
{
  "action": "log" | "delete" | "delete_after" | "status_month",
  "auto": bool,
  "start": "HH:MM" | null,
  "end": "HH:MM" | null,
  "datum": "DD.MM.YYYY" | null,
  "project_query": "kljucne reci projekta za pretragu" | null,
  "domain_hint": "IZVODJENJE" | "PROJEKTOVANJE" | "OPSTE" | "NADZOR" | "SERVIS" | "KONTROLA" | "BZR" | null,
  "comment": "verbatim komentar korisnika" | null
}

OPSTE:
- Ignoriši filere i okidace: "hej", "ej", "pčelo", "bee", "daj mi", "molim te", "ajde", "recimo", "neki", "glupi", "daj"
- Tekuća godina je 2026.

VREME:
- "od 8h do 13h" → start="08:00", end="13:00"
- "do 15" ili "dopuni do 15" → auto=true, start=null, end="15:00"
- Ako start > end (npr. od 18h do 12h) → verovatno greška govora: swap ih ako ima smisla, ili postavi start=null i auto=true
- "pola 9" = 08:30, "cetvrt do 5" = 16:45

PROJEKAT (project_query):
- Sve kljucne reci koje opisuju projekat — red reci NIJE vazan ("AI hub orion" = "orion AI hub")
- Ukloni SAMO vezne reci bez znacenja: "projekat", "u projektu", "na projektu", "domen", "u domenu"
- VAZNO: reci kao "opste", "servis", "nadzor", "izvođenje" mogu biti DEO NAZIVA projekta!
  Ako su deo naziva projekta — MORA ih biti u project_query.
  "dalibor opste" → project_query="dalibor opste" (opste je deo naziva projekta, ne samo domen)
  "dalibor servis" → project_query="dalibor servis"
- Sacuvaj specificne nazive, skracenice i brojeve

DOMEN (domain_hint) — OPCIONALAN filter, NEZAVISAN od project_query:
- domain_hint postavi SAMO kad korisnik eksplicitno kaze "u domenu X" ili "domen X"
- Ako korisnik kaze samo "dalibor opste" (bez "u domenu") → domain_hint=null, project_query="dalibor opste"
- "u domenu izvođenje" → IZVODJENJE
- "u domenu projektovanje" → PROJEKTOVANJE
- "u domenu opste" / "u domenu neradno" → OPSTE
- "u domenu nadzor" → NADZOR
- "u domenu servis" → SERVIS

KOMENTAR: prepiši verbatim sve što je korisnik rekao posle reči "komentar"

BRISANJE:
- "obrisi od X do Y" → action=delete, start=X, end=Y
- "obrisi sve posle X" → action=delete_after, start=X
- Datum: "za 25 februar" → "25.02.2026"

STATUS: "status za mesec" / "da li je sve uneseno" → action=status_month

Primeri:
"od 8 do 13 projekat resnik elektro komentar popravka" →
{"action":"log","auto":false,"start":"08:00","end":"13:00","datum":null,"project_query":"resnik elektro","domain_hint":null,"comment":"popravka"}

"ej pčelo upiši od 9h do 17h u domenu izvođenje projekat EMS Resnik komentar terenska merenja" →
{"action":"log","auto":false,"start":"09:00","end":"17:00","datum":null,"project_query":"EMS Resnik","domain_hint":"IZVODJENJE","comment":"terenska merenja"}

"dopuni do 16 projekat dalibor opste servis komentar mejlovi" →
{"action":"log","auto":true,"start":null,"end":"16:00","datum":null,"project_query":"dalibor opste servis","domain_hint":null,"comment":"mejlovi"}

"od 8 do 14 na projektu dalibor opste komentar tehnicka podrska" →
{"action":"log","auto":false,"start":"08:00","end":"14:00","datum":null,"project_query":"dalibor opste","domain_hint":null,"comment":"tehnicka podrska"}

"upiši od 9 do 17 u domenu izvođenje projekat EMS Resnik komentar merenja" →
{"action":"log","auto":false,"start":"09:00","end":"17:00","datum":null,"project_query":"EMS Resnik","domain_hint":"IZVODJENJE","comment":"merenja"}

"upiši od 18h do 12h u domenu izvođenje projekat orion AI hub komentar Idem na orion AI hub da pregovaram sa investitorom" →
{"action":"log","auto":false,"start":"08:00","end":"12:00","datum":null,"project_query":"orion AI hub","domain_hint":"IZVODJENJE","comment":"Idem na orion AI hub da pregovaram sa investitorom"}

"daj mi status za ovaj mesec" →
{"action":"status_month","auto":false,"start":null,"end":null,"datum":null,"project_query":null,"domain_hint":null,"comment":null}

"obrisi od 16 do 17" →
{"action":"delete","auto":false,"start":"16:00","end":"17:00","datum":null,"project_query":null,"domain_hint":null,"comment":null}

"obrisi sve unose posle 16 sati za dan 25 februar" →
{"action":"delete_after","auto":false,"start":"16:00","end":null,"datum":"25.02.2026","project_query":null,"domain_hint":null,"comment":null}
"""


# ── Chat sistem prompt ─────────────────────────────────────────────────────────

_CHAT_SYSTEM = """Ti si **Pčela** 🐝, AI asistent ugrađen u TVI Bee — aplikaciju za evidenciju radnog vremena kompanije Team Inoving.
Razgovaraš sa korisnikom na srpskom jeziku. Budi prijatan, koncizan i koristan. Koristi markdown za formatiranje.

## O TVI Bee aplikaciji
**TVI Bee** je web aplikacija i REST API koji omogućava korisnicima da evidentiraju radno vreme na TVI (Team Inoving) platformi direktno iz browsera ili Android aplikacije — bez da moraju da otvaraju sam TVI sajt.

### Arhitektura sistema
- **Backend**: Python 3.12 + FastAPI, teče na Debian Linux CT (Proxmox), port 7000
- **TVI konekcija**: DDP (Distributed Data Protocol) WebSocket klijent — isti protokol koji koristi Meteor.js framework; konekcija ide na `wss://tvi.meteor.teaminoving2.rs` preko SockJS
- **Autentifikacija**: HTTP Basic Auth → korisnik unosi TVI kredencijale u browser, backend ih validira DDP login-om i kešira sesiju 1 sat
- **AI**: Google Gemini 2.5 Flash — za parsiranje glasovnih komandi I ćaskanje (ti sama!)
- **Lokalna baza projekata**: SQLite fajl `projects.db` sa ~2931 projekata iz 7 domena, sinhronizovana sa TVI serverom (~3 min)
- **Logovanje**: JSONL fajl `logs/api.log` — svaki upis, brisanje, glasovni unos, auth greška

### Šta korisnik može da radi
1. **Evidencija radnog vremena** — unosi od/do vreme, bira projekat, dodaje komentar
2. **Glasovni unos** 🎤 — govori prirodnim jezikom ("od 8 do 14 na projektu Resnik komentar merenja"), Gemini parsira, Python fuzzy-matchuje projekat
3. **Pretraga projekata** — offline pretraga u lokalnoj SQLite bazi, tolerišu se greške u spelingu i crtice
4. **Istorija** — pregled po periodu, sortiran po danima
5. **Uređivanje/brisanje** — izmena ili brisanje zapisa direktno iz Danas pregleda
6. **Kilometraža (materijalni troškovi)** 🚗 — evidentiranje pređenih kilometara službeним automobilom:
   - Svaki zaposleni ima zadužen automobil koji se automatski bira pri otvaranju forme
   - Unosi se početna i krajnja kilometraža; sistem računa distancu i cenu (din/km)
   - Kilometraža se upisuje kao materijalni trošak na određeni projekat (obično "Ime Prezime Automobil Kuća-Posao" ili "Automobil Privatno")
   - Svi unosi se čuvaju lokalno u SQLite bazi i šalju na TVI server

### API endpointi
- `GET /api/status?datum=DD.MM.YYYY` — zapisi za dan (projekat, od-do, sati, komentar)
- `GET /api/day-with-ids` — isto + record ID-ovi (sporije ~5s zbog DDP subscription-a)
- `POST /api/log` — upis radnog vremena `{end_time, start_time?, comment?, project_number?, datum?}`
- `DELETE /api/record/{id}` — brisanje jednog zapisa
- `DELETE /api/day?datum=` — brisanje svih zapisa za dan
- `GET /api/history?od=&do=` — istorija perioda
- `GET /api/search?pojam=&domen=` — pretraga projekata (offline)
- `POST /api/parse-voice` — AI parsiranje glasovne komande (transkript → strukturisani JSON)
- `POST /api/sync` — sinhronizacija baze projekata sa TVI serverom (~3 min, samo admin)
- `GET /api/admin/logs?user=&action=&datum=` — server logovi (samo admin)
- `GET /api/me` — info o prijavljenom korisniku
- `GET /api/export` — Excel izveštaj za sve korisnike iz accounts.csv
- `GET /api/cars` — lista automobila sa zaduženim vozačima, ID korisnikovog auta, podrazumevani km projekti
- `GET /api/mileage?datum=` — kilometraže za dan
- `GET /api/mileage/history?od=&do=` — kilometraže za period (grupisane po danu, sa sumama)
- `POST /api/mileage` — upis kilometraže (car_id, start_km, end_km, activities_id, requests_id, date)

### Kako funkcioniše glasovni unos (korak po korak)
1. Korisnik klikne 🎤 u headeru
2. Browser Web Speech API snima i pretvara govor u tekst
3. Tekst se šalje na `/api/parse-voice`
4. Gemini 2.5 Flash ekstraktuje: akcija (log/delete/status), vreme (start/end), datum, project_query, domain_hint, komentar
5. Python fuzzy matching (`difflib.SequenceMatcher`) pretražuje SQLite bazu po project_query, uz filtriranje po domenu i godini
6. Vraća se top 5 kandidata — korisnik vidi preview sa dropdown-om za odabir projekta
7. Korisnik potvrđuje → POST /api/log se šalje na backend

### Baza projekata — detalji
- **7 domena**: Projektovanje (6), Izvođenje (8), Opšte i Neradno (9), Nadzor (10), Servis (12), Tehnička kontrola (13), BZR (14)
- Svaki projekat ima: activityNumber, name, domain, activities_id (OID), requests_id (OID)
- Fuzzy matching tolerišu greške u spelingu (>75% sličnost), deli po razmacima I crticama
- Automatski filtrira projekte iz prethodnih godina (traži `20XX` u nazivu)
- Sync traje ~3 minuta jer paginira sve domene stranicu po stranicu

### Šta se loguje i gde
Sve važne akcije se beleže u `/opt/tvi-bee/logs/api.log` (JSONL — jedan JSON per linija):
- **log (upis)**: korisnik, IP adresa, datum, od-do, projekat, komentar, record_id
- **parse-voice**: ceo transkript, AI parsiran rezultat (akcija, start, end, projekat query, komentar), matchovani projekat, top 3 kandidata
- **delete**: ko je obrisao koji record_id
- **delete-day**: ko je obrisao ceo dan, koliko zapisa
- **auth**: neuspešne prijave (korisnik, IP, opis greške)

### Višekorisnički sistem
- Svaki korisnik se prijavljuje sa **sopstvenim TVI korisničkim imenom i lozinkom** — nema posebnih Pčela naloga
- Backend validira kredencijale DDP login-om, kešira sesiju 1 sat po SHA256(username:password) ključu
- Svaki korisnik vidi samo svoje podatke
- Master/admin korisnik (dalibor.gmitrovic) ima pristup Admin panelu: logovi svih korisnika, MCP alati, sync projekata

### Saveti za efikasno korišćenje
- **"Dopuni do X"** — ako ne navodiš start, automatski nastavlja od kraja poslednjeg zapisa
- **Glasovni shortcuci**: "ej pčelo upiši od 9 do 17 projekat EMS Resnik komentar merenja"
- **Pretraga projekata**: piši deo naziva, aplikacija traži po svim rečima (AND logika)
- **Istorija**: filtriraj po periodu za pregled mesečnog rada
- **Brisanje**: u Danas pregledu klikni 🗑 ikonicu pored zapisa (učitava ID automatski u pozadini)

## Tvoje mogućnosti u ovom razgovoru
- Odgovaraj na pitanja o aplikaciji, funkcionalnostima, API-ju, arhitekturi
- Analiziraj korisnikove podatke koje vidiš u kontekstu (danas, mesec)
- Daj konkretne savete za korišćenje glasovnog unosa, pretrage, itd.
- Objasni tehničke detalje kad korisnik pita
- Pomozi pri rešavanju problema ("zašto nije pronašao projekat", "kako da editujem")
- **Automobili**: odgovaraj na pitanja o tome ko duži koji auto, koliko je pređeno km za dan/period, koliki su troškovi
- **Kilometraža**: prikaži detalje o pređenim kilometrima, troškovima, uporedi po danima/periodima
- Budi razgovorljiv i prijatan — nije samo pomoćnik, nego pravi AI asistent koji poznaje sistem
"""


def _build_user_context_sync(session: dict) -> str:
    """Gradi live kontekst korisnika za chat: ceo mesec (po danima), top projekti."""
    today = date.today()
    _day_start_ms = int(datetime(today.year, today.month, today.day).timestamp() * 1000)
    _, e_ms = _day_bounds(today)
    month_start = today.replace(day=1)
    ms_ms = int(datetime(month_start.year, month_start.month, 1).timestamp() * 1000)

    lines = [f"\n## Live podaci — {session['full_name']} ({session['username']})"]
    lines.append(f"Današnji datum: {today.strftime('%A, %d.%m.%Y')}")

    # Jedan DDP poziv za ceo mesec (today je podskup)
    try:
        month_recs = _fetch_records(session["user_id"], session["full_name"], ms_ms, e_ms, session)
    except Exception as ex:
        lines.append(f"Greška pri učitavanju podataka: {ex}")
        month_recs = []

    # Danas (filter iz meseca)
    today_recs = sorted(
        [r for r in month_recs if r["startTime"]["$date"] >= _day_start_ms],
        key=lambda x: x["startTime"]["$date"],
    )

    # IDs za danas — poseban DDP subscription poziv
    today_id_map: dict[int, str] = {}
    try:
        ddp_ids = MeteorDDP(_env("METEOR_WSS_URL"))
        if ddp_ids.connect(timeout=8):
            ddp_ids.login(session.get("username", _env("USERNAME")),
                          session.get("password", _env("PASSWORD")))
            id_recs = ddp_ids.get_request_time_ids_for_day(
                session.get("user_id", _env("USER_ID")), _day_start_ms, e_ms
            )
            ddp_ids.close()
            for ir in id_recs:
                today_id_map[ir["start_ms"]] = ir["id"]
    except Exception:
        pass

    if today_recs:
        today_h = round(sum(r["hours"] for r in today_recs), 2)
        lines.append(f"Danas: {len(today_recs)} zapisa, **{today_h}h**")
        for r in today_recs:
            st_ms = r["startTime"]["$date"]
            rid = today_id_map.get(st_ms, "")
            lines.append(
                f"  - [{rid}] {_ms_to_str(st_ms)}–{_ms_to_str(r['endTime']['$date'])}"
                f" | {r.get('requestName', '?')} | {r.get('comment', '') or '-'}"
            )
    else:
        lines.append("Danas: nema upisanih sati.")

    # Per-day mesec
    if month_recs:
        days: dict[str, list] = {}
        proj_h: dict[str, float] = {}
        for r in month_recs:
            dk = datetime.fromtimestamp(r["startTime"]["$date"] / 1000).strftime("%d.%m.")
            days.setdefault(dk, []).append(r)
            pn = r.get("requestName", "?")
            proj_h[pn] = proj_h.get(pn, 0.0) + r["hours"]

        month_hours = round(sum(r["hours"] for r in month_recs), 2)
        month_rsd   = round(sum(r["total"]  for r in month_recs))
        work_days   = sum(1 for v in days.values() if round(sum(x["hours"] for x in v), 2) > 0)
        lines.append(
            f"\n### Mesec {month_start.strftime('%m.%Y')}"
            f" ({month_start.strftime('%d.%m.')}–{today.strftime('%d.%m.')}): "
            f"{len(month_recs)} zapisa, **{month_hours}h**, {work_days} radnih dana, {month_rsd:,} RSD"
        )
        for dk in sorted(days.keys()):
            drecs = sorted(days[dk], key=lambda x: x["startTime"]["$date"])
            dh = round(sum(r["hours"] for r in drecs), 2)
            entries = "; ".join(
                f"[{_oid_value(r.get('_id',''))}] "
                f"{_ms_to_str(r['startTime']['$date'])}–{_ms_to_str(r['endTime']['$date'])}"
                f" {r.get('requestName','?')}"
                + (f" ({r.get('comment','')})" if r.get("comment") else "")
                for r in drecs
            )
            lines.append(f"  {dk}: {dh}h — {entries}")

        top5 = sorted(proj_h.items(), key=lambda x: -x[1])[:5]
        lines.append("\nTop projekti: " + ", ".join(f"{p} ({round(h,1)}h)" for p, h in top5))
    else:
        lines.append(f"\nOvaj mesec: nema zapisa.")

    try:
        conn = sqlite3.connect(PROJECTS_DB)
        proj_count = conn.execute("SELECT COUNT(*) FROM projects").fetchone()[0]
        conn.close()
        lines.append(f"Lokalna baza projekata: {proj_count} projekata")
    except Exception:
        pass

    # Automobil korisnika i kilometraža danas/mesec
    try:
        cars = _load_cars_from_db()
        my_car_id = _find_my_car_id(session.get("full_name", ""), cars)
        my_car = next((c for c in cars if c["_id"] == my_car_id), None) if my_car_id else None
        if my_car:
            lines.append(f"\n### Automobil")
            lines.append(f"Zadužen: **{my_car['name']}** ({my_car['unitPrice']} din/km)")

        _ensure_mileage_table()
        today_str = today.strftime("%Y-%m-%d")
        month_start_str = month_start.strftime("%Y-%m-%d")
        conn = sqlite3.connect(PROJECTS_DB)
        conn.row_factory = sqlite3.Row
        today_ml = conn.execute(
            "SELECT car_name, start_km, end_km, amount, total, project_name FROM mileage_log "
            "WHERE username = ? AND date = ? ORDER BY created_at",
            (session["username"], today_str),
        ).fetchall()
        month_ml = conn.execute(
            "SELECT SUM(amount) as km, SUM(total) as din, COUNT(*) as cnt FROM mileage_log "
            "WHERE username = ? AND date >= ? AND date <= ?",
            (session["username"], month_start_str, today_str),
        ).fetchone()
        conn.close()

        if today_ml:
            lines.append(f"Danas kilometraža: {len(today_ml)} unosa")
            for m in today_ml:
                lines.append(f"  - {m['car_name']}: {m['start_km']}→{m['end_km']} ({m['amount']} km, {m['total']} din) — {m['project_name'] or '?'}")
        if month_ml and month_ml["cnt"] and month_ml["cnt"] > 0:
            lines.append(f"Mesec kilometraža: {month_ml['cnt']} unosa, **{month_ml['km']} km**, {round(month_ml['din'])} din")
    except Exception:
        pass

    if session["username"] == MASTER_USER:
        lines.append("Uloga: **Admin** (master korisnik — pristup logovima, MCP, sync)")

    return "\n".join(lines)


_DOMAIN_MAP = {
    "IZVODJENJE":   ["IZVODJENJE"],
    "PROJEKTOVANJE": ["Projektovanje", "PROJEKTOVANJE"],
    "OPSTE":        ["OPSTE", "OPSTE_I_NERADNO"],
    "NADZOR":       ["NADZOR"],
    "SERVIS":       ["SERVIS"],
    "KONTROLA":     ["TEHNICKA_KONTROLA", "KONTROLA"],
    "BZR":          ["BZR_I_PPZ", "BZR"],
}


def _search_candidates(query: str, domain_hint: str = "", max_results: int = 5) -> list[dict]:
    """Pretraži sve projekte tekuce godine, skoruj po poklapanju reci (tolerise greske u spelingu).
    Ucitava sve projekte u Python i radi score — brzo za ~3000 redova."""
    if not PROJECTS_DB.exists() or not query.strip():
        return []

    def _norm(t: str) -> str:
        return ''.join(
            c for c in unicodedata.normalize('NFD', t.lower())
            if unicodedata.category(c) != 'Mn'
        )

    # Preskoci kratke reci (od, do, za, na...)
    words = [_norm(w) for w in query.split() if len(w) >= 3]
    if not words:
        return []

    current_year = str(date.today().year)

    try:
        conn = sqlite3.connect(PROJECTS_DB)
        rows = conn.execute(
            "SELECT activity_number, name, domain_name FROM projects "
            "ORDER BY domain_code+0, CAST(activity_number AS INTEGER) DESC"
        ).fetchall()
        conn.close()
    except Exception:
        return []

    def _wrong_year(name: str) -> bool:
        years = re.findall(r'\b(20\d\d)\b', name)
        return bool(years) and current_year not in years

    def _word_score(qw: str, norm_name: str) -> float:
        """Vrati skor poklapanja jedne reci upita sa nazivom projekta.
        2.0 = tacno poklapanje (substring), 0.75..1.0 = fuzzy (~75%), 0 = nema."""
        if qw in norm_name:
            return 2.0
        if len(qw) < 4:
            return 0.0  # kratke reci — samo tacno poklapanje
        # Razdvoji po razmacima I crticama (npr. "resnik-elektro" → ["resnik", "elektro"])
        for tw in re.split(r'[\s\-/]+', norm_name):
            if len(tw) < 4:
                continue
            r = difflib.SequenceMatcher(None, qw, tw).ratio()
            if r >= 0.75:
                return r
        return 0.0

    # Domain filter: ako je zadan domain_hint, dozvoli samo projekte tog domena
    allowed_domains = set()
    if domain_hint:
        hint_upper = domain_hint.upper()
        for key, variants in _DOMAIN_MAP.items():
            if key in hint_upper or hint_upper in key:
                allowed_domains.update(v.upper() for v in variants)

    scored = []
    for r in rows:
        if _wrong_year(r[1]):
            continue
        if allowed_domains and (r[2] or "").upper() not in allowed_domains:
            continue
        norm_name = _norm(r[1])
        score = sum(_word_score(w, norm_name) for w in words)
        if score > 0:
            scored.append((score, r))

    # Sortiraj: vise poklapanja = bolji, zatim noviji activity_number
    scored.sort(key=lambda x: (-x[0], -int(x[1][0] or 0)))

    return [{"activity_number": r[0] or "", "name": r[1], "domain": r[2], "score": sc}
            for sc, r in scored[:max_results]]


def _ai_parse(transcript: str) -> dict:
    """LLM izvlaci strukturu komande (akcija, vreme, project_query, komentar).
    Matchovanje projekta rade _search_candidates() i parse_voice endpoint."""
    normalized = _normalize_sr(transcript)
    today_str = date.today().strftime("%d.%m.%Y")
    user_text = f"Danas je {today_str}. {normalized}"
    backend = os.getenv("AI_BACKEND", "").strip().lower()

    if backend == "groq":
        api_key = os.getenv("GROQ_API_KEY", "").strip()
        if not api_key:
            raise RuntimeError("GROQ_API_KEY nije podesen u .env")
        model = os.getenv("GROQ_MODEL", "llama-3.1-8b-instant")
        payload = json.dumps({
            "model": model,
            "messages": [
                {"role": "system", "content": _VOICE_PROMPT},
                {"role": "user",   "content": user_text},
            ],
            "temperature": 0,
            "max_tokens": 150,
        }).encode()
        req = urllib.request.Request(
            "https://api.groq.com/openai/v1/chat/completions",
            data=payload,
            headers={"Authorization": f"Bearer {api_key}",
                     "Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            resp = json.loads(r.read())
        content = resp["choices"][0]["message"]["content"].strip()

    elif backend == "ollama":
        url   = os.getenv("OLLAMA_URL",   "http://localhost:11434")
        model = os.getenv("OLLAMA_MODEL", "llama3.2:3b")
        payload = json.dumps({
            "model":      model,
            "prompt":     f"{_VOICE_PROMPT}\n\nKomanda: {user_text}",
            "stream":     False,
            "format":     "json",
            "keep_alive": "10m",
        }).encode()
        req = urllib.request.Request(
            f"{url}/api/generate",
            data=payload,
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=120) as r:
            resp = json.loads(r.read())
        content = resp["response"].strip()

    elif backend == "gemini":
        api_key = os.getenv("GEMINI_API_KEY", "").strip()
        if not api_key:
            raise RuntimeError("GEMINI_API_KEY nije podesen u .env")
        model = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
        payload = json.dumps({
            "system_instruction": {"parts": [{"text": _VOICE_PROMPT}]},
            "contents": [{"parts": [{"text": user_text}]}],
            "generationConfig": {
                "temperature": 0,
                "maxOutputTokens": 2000,
            },
        }).encode()
        req = urllib.request.Request(
            f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}",
            data=payload,
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=30) as r:
            resp = json.loads(r.read())
        parts = resp.get("candidates", [{}])[0].get("content", {}).get("parts", [])
        # kod thinking modela (gemini-2.5-flash) uzmi poslednji deo koji sadrzi JSON
        text_parts = [p["text"] for p in parts if p.get("text", "").strip()]
        content = text_parts[-1].strip() if text_parts else ""

    else:
        raise ValueError(f"AI_BACKEND='{backend}' nije podrzan. Postavi 'gemini', 'groq' ili 'ollama'.")

    if "{" in content:
        content = content[content.index("{"):content.rindex("}")+1]
    return json.loads(content)


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/api/status")
async def status(
    datum: str = Query(default="", description="DD.MM.YYYY, prazno = danas"),
    session: dict = Depends(check_auth),
):
    """Prikazuje upisano radno vreme za dan."""
    def _run():
        user_id   = session["user_id"]
        user_name = session["full_name"]
        today = date.today()
        if datum:
            try:
                today = _parse_date(datum)
            except Exception:
                raise ValueError(f"Nevazan format datuma: '{datum}'. Koristiti DD.MM.YYYY.")
        s_ms, e_ms = _day_bounds(today)
        records = _fetch_records(user_id, user_name, s_ms, e_ms, session)
        result = {
            "datum": today.strftime("%d.%m.%Y"),
            "dan": WEEKDAYS[today.weekday()],
            "records": [],
            "total_hours": 0.0,
            "total_rsd": 0,
            "last_end": None,
        }
        if records:
            result["records"] = [
                _record_to_dict(r)
                for r in sorted(records, key=lambda x: x["startTime"]["$date"])
            ]
            result["total_hours"] = round(sum(r["hours"] for r in records), 4)
            result["total_rsd"] = round(sum(r["total"] for r in records))
            records_sorted = sorted(records, key=lambda x: x["endTime"]["$date"])
            result["last_end"] = _ms_to_str(records_sorted[-1]["endTime"]["$date"])
        return result

    try:
        return await asyncio.wait_for(asyncio.to_thread(_run), timeout=45.0)
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="TVI server ne odgovara (timeout). Pokusaj ponovo.")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/day-with-ids")
async def day_with_ids(
    datum: str = Query(default="", description="DD.MM.YYYY, prazno = danas"),
    session: dict = Depends(check_auth),
):
    """Status za dan sa record ID-ovima (sporo ~8s zbog subscription-a)."""
    def _run():
        user_id   = session["user_id"]
        user_name = session["full_name"]
        today = date.today()
        if datum:
            try:
                today = _parse_date(datum)
            except Exception:
                raise ValueError(f"Nevazan format datuma: '{datum}'. Koristiti DD.MM.YYYY.")
        s_ms, e_ms = _day_bounds(today)

        # Brzo: history podaci
        records = _fetch_records(user_id, user_name, s_ms, e_ms, session)

        # Sporo: IDs iz subscription-a
        ddp = _connect_login(session)
        id_records = ddp.get_request_time_ids_for_day(user_id, s_ms, e_ms)
        ddp.close()

        # Cross-reference po start_ms
        id_map: dict[int, str] = {}
        for r in id_records:
            if r["start_ms"] not in id_map:
                id_map[r["start_ms"]] = r["id"]

        result_records = []
        for r in sorted(records, key=lambda x: x["startTime"]["$date"]):
            rec = _record_to_dict(r)
            rec["id"] = id_map.get(r["startTime"]["$date"])
            result_records.append(rec)

        return {
            "datum": today.strftime("%d.%m.%Y"),
            "dan": WEEKDAYS[today.weekday()],
            "records": result_records,
            "total_hours": round(sum(r["hours"] for r in records), 4),
            "total_rsd": round(sum(r["total"] for r in records)),
            "last_end": _ms_to_str(
                max(r["endTime"]["$date"] for r in records)
            ) if records else None,
        }

    try:
        return await asyncio.wait_for(asyncio.to_thread(_run), timeout=45.0)
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="TVI server ne odgovara (timeout). Pokusaj ponovo.")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/log")
async def log(req: LogRequest, request: Request, session: dict = Depends(check_auth)):
    """Upisuje radno vreme. Ako start_time nije zadat, nastavlja od poslednjeg kraja."""
    def _run():
        user_id   = session["user_id"]
        user_name = session["full_name"]
        price_h   = int(_env("PRICE_PER_HOUR"))
        def_act   = _env("DEFAULT_ACTIVITIES_ID")
        def_req   = _env("DEFAULT_REQUESTS_ID")

        if req.datum:
            try:
                today = _parse_date(req.datum)
            except Exception:
                raise ValueError(f"Nevazan format datuma: '{req.datum}'. Koristiti DD.MM.YYYY.")
        else:
            today = date.today()

        # Projekat
        activities_id = def_act
        requests_id   = def_req
        project_info  = f"podrazumevani ({def_act[:8]}...)"

        if req.project_number:
            proj = _lookup_project_db(req.project_number)
            if proj:
                activities_id, requests_id, proj_name = proj
                if _is_other_user_project(proj_name, user_name):
                    raise ValueError(
                        f"Nije dozvoljeno upisivati vreme na projekat '{proj_name[:50]}' "
                        f"koji pripada drugom korisniku."
                    )
                project_info = f"#{req.project_number} {proj_name[:40]}"
            else:
                project_info = f"#{req.project_number} (nije nadjen u bazi, koristim podrazumevani)"

        # Vremena
        try:
            end_ms = _time_to_ms(req.end_time, today)
        except Exception:
            raise ValueError(f"Nevazan format end_time: '{req.end_time}'. Koristiti HH:MM.")

        s_ms, e_ms = _day_bounds(today)
        ddp = _connect_login(session)
        day_result = ddp.get_history(user_id=user_id, user_name=user_name,
                                     start_ms=s_ms, end_ms=e_ms)
        existing = (day_result.get("result") or []) if day_result else []

        if req.start_time:
            try:
                start_ms = _time_to_ms(req.start_time, today)
            except Exception:
                ddp.close()
                raise ValueError(f"Nevazan format start_time: '{req.start_time}'. Koristiti HH:MM.")
        else:
            # Auto-extend od poslednjeg kraja
            if not existing:
                ddp.close()
                raise ValueError("Nema zapisa za danas. Navedi start_time za prvi unos dana.")
            start_ms = max(r["endTime"]["$date"] for r in existing)

        if end_ms <= start_ms:
            ddp.close()
            raise ValueError(
                f"Kraj ({req.end_time}) mora biti posle pocetka ({_ms_to_str(start_ms)})."
            )

        # Provera preklapanja sa postojećim unosima za taj dan
        for r in existing:
            r_start = r["startTime"]["$date"]
            r_end   = r["endTime"]["$date"]
            if start_ms < r_end and end_ms > r_start:
                ddp.close()
                raise ValueError(
                    f"Preklapanje sa postojećim unosom "
                    f"{_ms_to_str(r_start)}–{_ms_to_str(r_end)} "
                    f"({r.get('requestName', '?')}). "
                    f"Novi unos {_ms_to_str(start_ms)}–{req.end_time} se preklapa."
                )

        hours = round((end_ms - start_ms) / 3_600_000, 4)
        total = round(hours * price_h)

        result = ddp.add_request_time(
            hours=hours,
            price_per_hour=price_h,
            comment=req.comment,
            engaged_user_id=user_id,
            start_ms=start_ms,
            end_ms=end_ms,
            activities_id=activities_id,
            requests_id=requests_id,
        )
        ddp.close()

        if result and "result" in result:
            record_id = _oid_value(result["result"].get("_id", {}))
            ip = _get_client_ip(request)
            _log_event(session["username"], ip, "log", {
                "datum":   today.strftime("%d.%m.%Y"),
                "start":   _ms_to_str(start_ms),
                "end":     req.end_time,
                "hours":   hours,
                "project": project_info,
                "comment": req.comment or "",
                "record_id": record_id,
            })
            return {
                "success": True,
                "datum": today.strftime("%d.%m.%Y"),
                "start": _ms_to_str(start_ms),
                "end": req.end_time,
                "hours": hours,
                "total": total,
                "project": project_info,
                "comment": req.comment or "",
                "record_id": record_id,
            }
        elif result and "error" in result:
            raise RuntimeError(f"Greska sa servera: {result['error']}")
        else:
            raise RuntimeError("Nije dobijen odgovor od servera.")

    ip = _get_client_ip(request)
    try:
        return await asyncio.wait_for(asyncio.to_thread(_run), timeout=45.0)
    except asyncio.TimeoutError:
        _log_event(session["username"], ip, "log", {"end_time": req.end_time}, "504")
        raise HTTPException(status_code=504, detail="TVI server ne odgovara (timeout). Pokusaj ponovo.")
    except ValueError as e:
        _log_event(session["username"], ip, "log", {"end_time": req.end_time, "error": str(e)}, "400")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        _log_event(session["username"], ip, "log", {"end_time": req.end_time, "error": str(e)}, "500")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/record/{record_id}")
async def delete_record(record_id: str, request: Request, session: dict = Depends(check_auth)):
    """Brise jedan zapis po ID-u."""
    def _run():
        ddp = _connect_login(session)
        result = ddp.remove_request_time(record_id)
        ddp.close()
        if result is None:
            raise RuntimeError("Nije dobijen odgovor od servera.")
        if "error" in result:
            err = result["error"]
            if isinstance(err, dict) and err.get("error") == "item_already_deleted":
                raise ValueError(f"Zapis {record_id} ne postoji ili je vec obrisan.")
            raise RuntimeError(f"Greska sa servera: {err}")
        ip = _get_client_ip(request)
        _log_event(session["username"], ip, "delete", {"record_id": record_id})
        return {"success": True, "record_id": record_id,
                "message": f"Zapis {record_id} uspesno obrisan."}

    ip = _get_client_ip(request)
    try:
        return await asyncio.to_thread(_run)
    except ValueError as e:
        _log_event(session["username"], ip, "delete", {"record_id": record_id, "error": str(e)}, "404")
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        _log_event(session["username"], ip, "delete", {"record_id": record_id, "error": str(e)}, "500")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/day")
async def delete_day(
    datum: str = Query(default="", description="DD.MM.YYYY, prazno = danas"),
    request: Request = None,
    session: dict = Depends(check_auth),
):
    """Brise sve zapise za dan."""
    def _run():
        user_id = session["user_id"]
        if datum:
            try:
                day = _parse_date(datum)
            except Exception:
                raise ValueError(f"Nevazan format datuma: '{datum}'. Koristiti DD.MM.YYYY.")
        else:
            day = date.today()

        s_ms, e_ms = _day_bounds(day)
        ddp = _connect_login(session)
        records = ddp.get_request_time_ids_for_day(user_id, s_ms, e_ms)

        if not records:
            ddp.close()
            return {"datum": day.strftime("%d.%m.%Y"), "deleted": 0, "results": []}

        results = []
        for r in records:
            result = ddp.remove_request_time(r["id"])
            s = _ms_to_str(r["start_ms"])
            e = _ms_to_str(r["end_ms"])
            if result and "error" not in result:
                results.append({"id": r["id"], "start": s, "end": e, "success": True})
            else:
                err = (result or {}).get("error", {})
                results.append({"id": r["id"], "start": s, "end": e,
                                "success": False, "error": str(err)})
        ddp.close()

        deleted = sum(1 for r in results if r["success"])
        ip = _get_client_ip(request) if request else "?"
        _log_event(session["username"], ip, "delete-day", {
            "datum": day.strftime("%d.%m.%Y"), "deleted": deleted,
        })
        return {"datum": day.strftime("%d.%m.%Y"), "deleted": deleted, "results": results}

    ip = _get_client_ip(request) if request else "?"
    try:
        return await asyncio.wait_for(asyncio.to_thread(_run), timeout=45.0)
    except asyncio.TimeoutError:
        _log_event(session["username"], ip, "delete-day", {"datum": datum, "error": "timeout"}, "504")
        raise HTTPException(status_code=504, detail="TVI server ne odgovara (timeout). Pokusaj ponovo.")
    except ValueError as e:
        _log_event(session["username"], ip, "delete-day", {"datum": datum, "error": str(e)}, "400")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        _log_event(session["username"], ip, "delete-day", {"datum": datum, "error": str(e)}, "500")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/history")
async def history(
    od: str = Query(default="", description="DD.MM.YYYY, prazno = prvi dan meseca"),
    do: str = Query(default="", description="DD.MM.YYYY, prazno = danas"),
    session: dict = Depends(check_auth),
):
    """Prikazuje istoriju radnog vremena za period."""
    def _run():
        user_id   = session["user_id"]
        user_name = session["full_name"]
        try:
            s_ms, e_ms, sd, ed = _period_bounds(od, do)
        except Exception:
            raise ValueError("Nevazan format datuma. Koristiti DD.MM.YYYY.")

        records = _fetch_records(user_id, user_name, s_ms, e_ms, session)

        by_day: dict[date, list] = {}
        for r in records:
            day = datetime.fromtimestamp(r["date"]["$date"] / 1000).date()
            by_day.setdefault(day, []).append(r)

        days = []
        total_hours = 0.0
        total_rsd = 0
        for day in sorted(by_day.keys()):
            day_records = sorted(by_day[day], key=lambda x: x["startTime"]["$date"])
            dh = sum(r["hours"] for r in day_records)
            dr = sum(r["total"] for r in day_records)
            total_hours += dh
            total_rsd += dr
            days.append({
                "datum": day.strftime("%d.%m.%Y"),
                "dan": WEEKDAYS[day.weekday()],
                "hours": round(dh, 4),
                "total": round(dr),
                "records": [_record_to_dict(r) for r in day_records],
            })

        return {
            "od": sd.strftime("%d.%m.%Y"),
            "do": ed.strftime("%d.%m.%Y"),
            "radnih_dana": len(by_day),
            "total_hours": round(total_hours, 4),
            "total_rsd": round(total_rsd),
            "days": days,
        }

    try:
        return await asyncio.wait_for(asyncio.to_thread(_run), timeout=45.0)
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="TVI server ne odgovara (timeout). Pokusaj ponovo.")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/search")
async def search(
    pojam: str = Query(..., description="Deo naziva projekta"),
    domen: str = Query(default="", description="Kod domena: 6=Projektovanje, 8=Izvodjenje, "
                       "9=Opste i Neradno, 10=Nadzor, 12=Servis, 13=Tehnicka Kontrola, 14=BZR"),
    session: dict = Depends(check_auth),
):
    """Pretrazuje projekte u lokalnoj SQLite bazi (bez konekcije na server)."""
    def _run():
        if not PROJECTS_DB.exists():
            raise RuntimeError(
                f"Baza projekata ne postoji ({PROJECTS_DB}). "
                "Pokreni POST /api/sync za preuzimanje projekata."
            )
        # Normalizuj query i nazive (ukloni dijakritike) za robustno matchovanje
        def _norm(t: str) -> str:
            return ''.join(
                c for c in unicodedata.normalize('NFD', t.lower())
                if unicodedata.category(c) != 'Mn'
            )

        words_orig = [w.strip() for w in pojam.split() if w.strip()]
        norm_words = [_norm(w) for w in words_orig]

        # SQL: first-pass filter po prvoj riječi (brzo), Python: AND svih normalizovanih riječi
        conn = sqlite3.connect(PROJECTS_DB)
        if domen:
            rows = conn.execute(
                "SELECT activity_number, name, domain_name, id, requests_id "
                "FROM projects WHERE name LIKE ? AND domain_code = ? "
                "ORDER BY domain_code+0, CAST(activity_number AS INTEGER) DESC",
                (f"%{words_orig[0]}%", str(domen))
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT activity_number, name, domain_name, id, requests_id "
                "FROM projects WHERE name LIKE ? "
                "ORDER BY domain_code+0, CAST(activity_number AS INTEGER) DESC",
                (f"%{words_orig[0]}%",)
            ).fetchall()
        # Python-level: sve normalizovane riječi moraju biti u normalizovanom nazivu
        rows = [r for r in rows if all(nw in _norm(r[1]) for nw in norm_words)]

        # Filtriraj projekte iz prethodnih godina — prikazuj samo tekuću godinu (ili bez godine)
        current_year = str(date.today().year)
        def _wrong_year(name: str) -> bool:
            years = re.findall(r'\b(20\d\d)\b', name)
            return bool(years) and current_year not in years
        rows = [r for r in rows if not _wrong_year(r[1])]
        rows = [r for r in rows if not _is_other_user_project(r[1], session["full_name"])]
        conn.close()

        projects = [
            {
                "activity_number": num or "",
                "name": name,
                "domain": domain_name,
                "activities_id": act_id,
                "requests_id": req_id,
            }
            for num, name, domain_name, act_id, req_id in rows
        ]
        return {"pojam": pojam, "count": len(projects), "projects": projects}

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


SYNC_DOMAINS = [
    ("6",  "Projektovanje"),
    ("8",  "IZVODJENJE"),
    ("9",  "OPSTE_I_NERADNO"),
    ("10", "NADZOR"),
    ("12", "SERVIS"),
    ("13", "TEHNICKA_KONTROLA"),
    ("14", "BZR_I_PPZ"),
]


def _sync_projects_blocking(session: dict | None = None,
                             progress_cb=None) -> dict:
    """Sinhrona sinhronizacija projekata. Može se koristiti iz threada.

    session: dict sa username/password, ili None za env kredencijale.
    progress_cb: opcionalna f(msg: dict) za real-time napredak (SSE/log).
    Vraća {"success": True, "fetched_at": ..., "total": ..., "domains": {...}}.
    """
    PAGE_SIZE = 20
    PAUSE_SEC = 0.4
    OUTPUT_DIR = BASE_DIR / "projects"
    DB_PATH = OUTPUT_DIR / "projects.db"
    total_domains = len(SYNC_DOMAINS)

    def _cb(msg: dict):
        if progress_cb:
            progress_cb(msg)

    OUTPUT_DIR.mkdir(exist_ok=True)
    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id TEXT PRIMARY KEY, activity_number TEXT, name TEXT,
            domain_code TEXT, domain_name TEXT, requests_id TEXT, fetched_at TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_domain ON projects(domain_code)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_number ON projects(activity_number)")
    conn.commit()

    totals = {}
    for domain_idx, (dc, dn) in enumerate(SYNC_DOMAINS):
        pct_start = int(domain_idx / total_domains * 100)
        _cb({"stage": "domain_start", "domain": dn,
             "domain_idx": domain_idx + 1, "domain_total": total_domains,
             "pct": pct_start})

        ddp = None
        for attempt in range(3):
            try:
                ddp = _connect_login(session)
                break
            except Exception:
                if attempt < 2:
                    time.sleep(3)

        if ddp is None:
            _cb({"stage": "domain_error", "domain": dn,
                 "domain_idx": domain_idx + 1, "domain_total": total_domains})
            totals[dn] = 0
            time.sleep(2)
            continue

        all_docs = []
        page = 1
        while True:
            docs = ddp.search_activities_page(dc, "*", page=page, page_size=PAGE_SIZE)
            all_docs.extend(docs)
            _cb({"stage": "page", "domain": dn, "page": page,
                 "fetched": len(all_docs), "pct": pct_start})
            if len(docs) < PAGE_SIZE:
                break
            page += 1
            time.sleep(PAUSE_SEC)

        ddp.close()
        time.sleep(1)

        rows = []
        for doc in all_docs:
            req_id = _oid_value(doc.get("requests_id", ""))
            rows.append((
                doc["_id"], str(doc.get("activityNumber", "")),
                doc.get("name", ""), dc, dn, req_id, fetched_at,
            ))
        conn.executemany(
            "INSERT OR REPLACE INTO projects "
            "(id, activity_number, name, domain_code, domain_name, requests_id, fetched_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            rows,
        )
        conn.commit()
        totals[dn] = len(all_docs)

        pct_end = int((domain_idx + 1) / total_domains * 100)
        _cb({"stage": "domain_done", "domain": dn,
             "domain_idx": domain_idx + 1, "domain_total": total_domains,
             "count": len(all_docs), "pct": pct_end})
        time.sleep(PAUSE_SEC)

    conn.close()
    total = sum(totals.values())
    _cb({"stage": "finished", "total": total, "domains": totals,
         "fetched_at": fetched_at})
    return {"success": True, "fetched_at": fetched_at,
            "db_path": str(DB_PATH), "total": total, "domains": totals}


# ── Sync automobila (items) u SQLite ──────────────────────────────────────────

CARS_DB = BASE_DIR / "projects" / "projects.db"

# Zaduženja vozila iz putnih naloga (registracija -> vozač)
CAR_DRIVERS: dict[str, str] = {
    "BG 1039-FG": "Јован Ђорђевић",
    "BG 1165-EZ": "Бранко Станковић",
    "BG 1369-FC": "Коста Петровић",
    "BG 1392-OL": "Мићо Јарић",
    "BG 1456-LB": "Лазар Недељковић",
    "BG 1534-GT": "Марко Бранковић",
    "BG 1553-AS": "Марија Спасојевић",
    "BG 1553-CI": "Драган Вујић",
    "BG 1639-TT": "Марко Бурмазевић",
    "BG 1755-HT": "Далибор Гмитровић",
    "BG 1799-FN": "Милош Јовановић",
    "BG 1936-LI": "Немања Росић",
    "BG 2183-CD": "Немања Миросавић",
    "BG 2236-AJ": "Александра Гачевић",
    "BG 2412-TN": "Зоран Петровић",
    "BG 2423-KV": "Маријан Ђеровић",
    "BG 2497-VC": "Дејан Дељанин",
    "BG 2512-CA": "Пеђа Крстић",
    "BG 2537-JR": "Ђорђе Виријевић",
    "BG 2687-VF": "Милош Грбовић",
    "BG 2720-KL": "Милан Љубојевић",
    "BG 2741-KT": "Бојан Дошлов",
    "BG 2752-VD": "Данијела Лазаревић",
    "BG 2752-ZV": "Бојан Јовановић",
    "BG 2774-LR": "Момчило Здравковић",
    "BG 2909-BU": "Милош Милошевић",
    "BG 3106-AF": "Марко Икић",
    "BG 3156-AZ": "Радомир Керкез",
    "BG 3218-ZX": "Марко Ђуричић",
    "BG 865-BI": "Александар Мијаиловић",
    "BG 883-PO": "Бошко Вуковић",
}


_CYR_TO_LAT = str.maketrans(
    "АБВГДЕЖЗИЈКЛМНОПРСТУФХЦЧШабвгдежзијклмнопрстуфхцчш",
    "ABVGDEŽZIJKLMNOPRSTUFHCČŠabvgdežzijklmnoprstufhcčš",
)
_CYR_DIGRAPHS = [("Љ", "Lj"), ("Њ", "Nj"), ("Џ", "Dž"), ("Ђ", "Đ"), ("Ћ", "Ć"),
                  ("љ", "lj"), ("њ", "nj"), ("џ", "dž"), ("ђ", "đ"), ("ћ", "ć")]


def _cyr_to_lat(text: str) -> str:
    """Srpska ćirilica -> latinica transliteracija."""
    for cyr, lat in _CYR_DIGRAPHS:
        text = text.replace(cyr, lat)
    return text.translate(_CYR_TO_LAT)


def _match_driver(car_name: str) -> str:
    """Pronalazi vozača za auto na osnovu registracije u nazivu."""
    name_flat = car_name.replace(" ", "").replace("-", "").upper()
    for reg, driver in CAR_DRIVERS.items():
        reg_flat = reg.replace(" ", "").replace("-", "").upper()
        if reg_flat in name_flat:
            return driver
    return ""


def _find_my_car_id(full_name: str, cars: list[dict]) -> str | None:
    """Pronalazi ID automobila zaduženog za korisnika po imenu."""
    if not full_name:
        return None
    fn_norm = unicodedata.normalize("NFD", full_name.lower())
    fn_clean = "".join(c for c in fn_norm if unicodedata.category(c) != "Mn")
    for car in cars:
        driver_cyr = car.get("driver", "")
        if not driver_cyr:
            continue
        driver_lat = _cyr_to_lat(driver_cyr).lower()
        dl_norm = unicodedata.normalize("NFD", driver_lat)
        dl_clean = "".join(c for c in dl_norm if unicodedata.category(c) != "Mn")
        if dl_clean == fn_clean:
            return car["_id"]
    return None


def _sync_cars_blocking(session: dict | None = None) -> dict:
    """Povlači sve items (automobile) iz TVI-ja i čuva u SQLite."""
    ddp = _connect_login(session)
    try:
        docs = ddp.search_items(timeout=30)
    finally:
        ddp.close()

    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(CARS_DB)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cars (
            id TEXT PRIMARY KEY, code TEXT, name TEXT, type TEXT,
            base_unit TEXT, unit_price REAL, driver TEXT DEFAULT '',
            fetched_at TEXT
        )
    """)
    # Migracija: dodaj driver kolonu ako ne postoji
    cols = [r[1] for r in conn.execute("PRAGMA table_info(cars)").fetchall()]
    if "driver" not in cols:
        conn.execute("ALTER TABLE cars ADD COLUMN driver TEXT DEFAULT ''")
    conn.commit()

    rows = []
    for doc in docs:
        fields = doc if "code" in doc else doc.get("fields", doc)
        up = fields.get("unitPrice")
        if not up:
            continue
        car_name = fields.get("name", "")
        rows.append((
            doc.get("_id", doc.get("id", "")),
            str(fields.get("code", "")),
            car_name,
            str(fields.get("type", "")),
            fields.get("baseUnitOfMeasure", "kilometer"),
            float(up),
            _match_driver(car_name),
            fetched_at,
        ))

    conn.execute("DELETE FROM cars")
    conn.executemany(
        "INSERT INTO cars (id, code, name, type, base_unit, unit_price, driver, fetched_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        rows,
    )
    conn.commit()
    conn.close()
    return {"count": len(rows), "fetched_at": fetched_at}


def _sync_timesheets_blocking(
    od: str = "",
    do: str = "",
    progress_cb=None,
) -> dict:
    """Povlači radno vreme svih zaposlenih iz accounts.csv i čuva u SQLite.

    od: Početak perioda DD.MM.YYYY (default: 01.01.2020)
    do: Kraj perioda DD.MM.YYYY (default: danas)
    progress_cb: opcionalna f(msg: dict) za real-time napredak
    """
    if not ACCOUNTS_CSV.exists():
        raise RuntimeError(f"Fajl '{ACCOUNTS_CSV}' ne postoji.")
    accounts = []
    with open(ACCOUNTS_CSV, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            accounts.append({
                "username":       row["username"].strip(),
                "password":       row["password"].strip(),
                "user_id":        row["user_id"].strip(),
                "full_name":      row["full_name"].strip(),
                "price_per_hour": int(row.get("price_per_hour", "2300").strip() or "2300"),
            })
    if not accounts:
        raise RuntimeError("accounts.csv je prazan.")

    s_date = datetime.strptime(od, "%d.%m.%Y").date() if od else date(2020, 1, 1)
    e_date = datetime.strptime(do, "%d.%m.%Y").date() if do else date.today()
    s_ms = int(datetime(s_date.year, s_date.month, s_date.day).timestamp() * 1000)
    e_ms = int(datetime(e_date.year, e_date.month, e_date.day, 23, 59, 59).timestamp() * 1000)

    meteor_url = _env("METEOR_WSS_URL")
    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(PROJECTS_DB)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS request_times (
            id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            user_name TEXT NOT NULL,
            start_ms INTEGER,
            end_ms INTEGER,
            date_ms INTEGER,
            hours REAL,
            price_per_hour REAL,
            total REAL,
            comment TEXT,
            project_name TEXT,
            fetched_at TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_rt_user ON request_times(user_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_rt_date ON request_times(date_ms)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            user_id TEXT PRIMARY KEY,
            username TEXT,
            full_name TEXT,
            price_per_hour INTEGER,
            last_synced_at TEXT
        )
    """)
    conn.commit()

    def _cb(msg: dict):
        if progress_cb:
            progress_cb(msg)

    total_records = 0
    synced_users = 0
    errors = []

    for i, acc in enumerate(accounts, 1):
        _cb({"type": "user", "user": acc["full_name"], "n": i, "total": len(accounts)})
        try:
            ddp = MeteorDDP(meteor_url)
            if not ddp.connect(timeout=15):
                errors.append(f"{acc['full_name']}: konekcija nije uspela")
                _cb({"type": "error", "user": acc["full_name"], "msg": "konekcija nije uspela"})
                continue
            if not ddp.login(acc["username"], acc["password"]):
                ddp.close()
                errors.append(f"{acc['full_name']}: prijava nije uspela")
                _cb({"type": "error", "user": acc["full_name"], "msg": "prijava nije uspela"})
                continue

            # user_id iz CSV-a ili iz login odgovora (za nove korisnike bez user_id)
            effective_user_id = acc["user_id"] or ddp.user_id or ""

            result = ddp.get_history(
                user_id=effective_user_id,
                user_name=acc["full_name"],
                start_ms=s_ms,
                end_ms=e_ms,
            )
            ddp.close()

            recs = result["result"] if result and "result" in result and result["result"] else []

            for r in recs:
                raw_id = r.get("_id")
                rec_id = raw_id.get("$value") if isinstance(raw_id, dict) else (str(raw_id) if raw_id else None)
                if not rec_id:
                    continue
                st = r.get("startTime")
                et = r.get("endTime")
                dt = r.get("date")
                conn.execute("""
                    INSERT OR REPLACE INTO request_times
                    (id, user_id, user_name, start_ms, end_ms, date_ms,
                     hours, price_per_hour, total, comment, project_name, fetched_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    rec_id,
                    effective_user_id,
                    acc["full_name"],
                    st["$date"] if isinstance(st, dict) else 0,
                    et["$date"] if isinstance(et, dict) else 0,
                    dt["$date"] if isinstance(dt, dict) else 0,
                    r.get("hours") or 0.0,
                    r.get("pricePerHour") or acc["price_per_hour"],
                    r.get("total") or 0.0,
                    r.get("comment") or "",
                    r.get("requestName") or "",
                    fetched_at,
                ))

            conn.execute("""
                INSERT OR REPLACE INTO employees (user_id, username, full_name, price_per_hour, last_synced_at)
                VALUES (?, ?, ?, ?, ?)
            """, (effective_user_id, acc["username"], acc["full_name"], acc["price_per_hour"], fetched_at))
            conn.commit()

            total_records += len(recs)
            synced_users += 1
            _cb({"type": "done", "user": acc["full_name"], "records": len(recs)})

        except Exception as e:
            errors.append(f"{acc['full_name']}: {e}")
            _cb({"type": "error", "user": acc["full_name"], "msg": str(e)})

        time.sleep(0.5)

    conn.close()
    return {
        "synced_users": synced_users,
        "total_records": total_records,
        "errors": errors,
        "period": f"{s_date:%d.%m.%Y} — {e_date:%d.%m.%Y}",
        "fetched_at": fetched_at,
    }


# ── Auto-sync scheduler (svake noći u 04:00) ─────────────────────────────────

def _auto_sync_run() -> None:
    """Pokreće sinhronizaciju projekata, automobila i radnog vremena."""
    _log_event("system", "localhost", "auto-sync", {"detail": "pokretanje"})
    try:
        result = _sync_projects_blocking(session=None)
        _write_last_sync(result["total"], result["domains"], mode="auto")
    except Exception as e:
        _log_event("system", "localhost", "auto-sync",
                   {"error": str(e)}, status="error")
    try:
        cars_result = _sync_cars_blocking(session=None)
        _log_event("system", "localhost", "auto-sync-cars",
                   {"count": cars_result["count"]})
    except Exception as e:
        _log_event("system", "localhost", "auto-sync-cars",
                   {"error": str(e)}, status="error")
    # Inkrementalni sync radnog vremena — poslednih 35 dana
    try:
        od_inc = (date.today() - timedelta(days=35)).strftime("%d.%m.%Y")
        ts_result = _sync_timesheets_blocking(od=od_inc)
        _log_event("system", "localhost", "auto-sync-timesheets",
                   {"records": ts_result["total_records"], "users": ts_result["synced_users"]})
    except Exception as e:
        _log_event("system", "localhost", "auto-sync-timesheets",
                   {"error": str(e)}, status="error")


def _scheduler_loop() -> None:
    """Daemon thread: spava do 04:00 i pokreće auto-sync."""
    while True:
        now = datetime.now()
        target = now.replace(hour=4, minute=0, second=0, microsecond=0)
        if target <= now:
            target += timedelta(days=1)
        sleep_sec = (target - now).total_seconds()
        _log_event("system", "localhost", "scheduler",
                   {"next_sync": target.strftime("%Y-%m-%d %H:%M:%S"),
                    "sleep_min": round(sleep_sec / 60)})
        time.sleep(sleep_sec)
        _auto_sync_run()


threading.Thread(target=_scheduler_loop, daemon=True, name="sync-scheduler").start()


@app.post("/api/sync")
async def sync(session: dict = Depends(check_auth)):
    """Osvezava lokalnu bazu projekata sa TVI servera (moze trajati 5-10 min)."""
    if session["username"] != MASTER_USER:
        raise HTTPException(403, "Samo admin može pokrenuti sinhronizaciju.")

    def _run():
        result = _sync_projects_blocking(session)
        _write_last_sync(result["total"], result["domains"], mode="manual")
        return result

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/sync/stream")
async def sync_stream(session: dict = Depends(check_auth)):
    """SSE stream sinhronizacije — šalje napredak u realnom vremenu."""
    if session["username"] != MASTER_USER:
        raise HTTPException(403, "Samo admin može pokrenuti sinhronizaciju.")

    q: queue.Queue = queue.Queue()

    def _run():
        try:
            result = _sync_projects_blocking(session, progress_cb=q.put)
            _write_last_sync(result["total"], result["domains"], mode="manual")
        except Exception as e:
            q.put({"stage": "error", "detail": str(e)})

    threading.Thread(target=_run, daemon=True).start()

    async def generate():
        while True:
            try:
                msg = q.get_nowait()
            except queue.Empty:
                await asyncio.sleep(0.1)
                continue
            yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
            if msg.get("stage") in ("finished", "error"):
                break

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/admin/sync/status")
async def sync_status(session: dict = Depends(check_auth)):
    """Vraća info o poslednjoj sinhronizaciji (vreme, ukupno projekata, mod)."""
    if session["username"] != MASTER_USER:
        raise HTTPException(403, "Samo admin.")
    data = _read_last_sync()
    if data is None:
        return {"last_sync": None}
    # Izračunaj sledeći auto-sync
    now = datetime.now()
    target = now.replace(hour=4, minute=0, second=0, microsecond=0)
    if target <= now:
        target += timedelta(days=1)
    return {
        "last_sync": data,
        "next_auto_sync": target.strftime("%Y-%m-%d %H:%M"),
    }


@app.post("/api/parse-voice")
async def parse_voice(req: ParseVoiceRequest, request: Request, session: dict = Depends(check_auth)):
    """AI parser glasovnih komandi. Zahtijeva AI_BACKEND=gemini|groq|ollama u .env"""
    ip = _get_client_ip(request)
    def _run():
        # LLM: izvuci strukturu (akcija, vreme, project_query, komentar)
        result = _ai_parse(req.transcript)

        # Python: matchuj projekat iz DB po project_query + domain_hint
        project_query = (result.get("project_query") or "").strip()
        domain_hint   = (result.get("domain_hint") or "").strip()
        candidates    = []
        act_num       = ""
        proj_name     = ""
        if project_query:
            candidates = _search_candidates(project_query, domain_hint=domain_hint)
            if candidates:
                best      = candidates[0]
                act_num   = best["activity_number"]
                proj_name = best["name"]

        response = {
            "action":          result.get("action", "log"),
            "auto":            bool(result.get("auto", True)),
            "start":           result.get("start") or None,
            "end":             str(result.get("end", "") or ""),
            "datum":           (result.get("datum") or "").strip() or None,
            "activity_number": act_num,
            "project_name":    proj_name,
            "project_query":   project_query if not act_num else "",
            "comment":         result.get("comment") or "",
            "candidates":      candidates,
        }
        _log_event(session["username"], ip, "parse-voice", {
            "transcript":    req.transcript,
            "ai_action":     response["action"],
            "ai_start":      response["start"],
            "ai_end":        response["end"],
            "ai_datum":      response["datum"],
            "project_query": project_query,
            "domain_hint":   domain_hint,
            "ai_comment":    response["comment"],
            "matched_num":   act_num,
            "matched_name":  proj_name,
            "candidates":    [f"#{c['activity_number']} {c['name']}" for c in candidates[:3]],
        })
        return response
    try:
        return await asyncio.to_thread(_run)
    except ValueError as e:
        _log_event(session["username"], ip, "parse-voice", {"transcript": req.transcript, "error": str(e)}, "400")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        _log_event(session["username"], ip, "parse-voice", {"transcript": req.transcript, "error": str(e)}, "500")
        raise HTTPException(status_code=500, detail=f"AI parser greška: {e}")


# ── Chat (ćaskanje sa Pčelom) ─────────────────────────────────────────────────
#
# Akcije se prenose kroz tekst — Gemini ugradi JSON tag na kraju odgovora,
# backend ga parsira i izvršava MCP alat (bez Gemini function calling API-ja).
#
def _chat_search_projects(pojam: str, domen: str = "") -> str:
    """Pretraži lokalne projekte za chat AI.

    Sve filtriranje radi u Pythonu (unicode normalizacija, AND po rečima).
    Zaobilazi SQLite LIKE koji ne podržava case-insensitive za Unicode (š, ć, đ...).
    """
    if not PROJECTS_DB.exists():
        return "Baza projekata ne postoji. Pokreni sinhronizaciju."
    if not pojam.strip():
        return "Upiši deo naziva projekta."

    def _norm(t: str) -> str:
        return ''.join(
            c for c in unicodedata.normalize('NFD', t.lower())
            if unicodedata.category(c) != 'Mn'
        )

    words = [w for w in pojam.split() if len(w) >= 2]
    if not words:
        return "Previše kratak pojam za pretragu."
    norm_words = [_norm(w) for w in words]

    conn = sqlite3.connect(PROJECTS_DB)
    try:
        if domen:
            rows = conn.execute(
                "SELECT activity_number, name, domain_name, id, requests_id "
                "FROM projects WHERE domain_code = ? "
                "ORDER BY domain_code+0, CAST(activity_number AS INTEGER) DESC",
                (str(domen),)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT activity_number, name, domain_name, id, requests_id "
                "FROM projects "
                "ORDER BY domain_code+0, CAST(activity_number AS INTEGER) DESC"
            ).fetchall()
    finally:
        conn.close()

    # Python filtriranje: sve normalizovane reči moraju biti u normalizovanom nazivu
    rows = [r for r in rows if all(nw in _norm(r[1]) for nw in norm_words)]

    # Samo tekuća i bezgodišnji projekti
    current_year = str(date.today().year)
    def _wrong_year(name: str) -> bool:
        years = re.findall(r'\b(20\d\d)\b', name)
        return bool(years) and current_year not in years
    rows = [r for r in rows if not _wrong_year(r[1])]

    if not rows:
        return f"Nema rezultata za '{pojam}'."

    lines = [f"Pronađeno {len(rows)} projekata za '{pojam}':\n"]
    cur_domain = None
    for num, name, domain_name, act_id, req_id in rows[:20]:
        if domain_name != cur_domain:
            cur_domain = domain_name
            lines.append(f"\n[{domain_name}]")
        lines.append(f"  project_number={num} → {name}")
    lines.append("\n⚠ KRITIČNO: U tvi_log pozivu, project_number mora biti TAČNO prepisana cifra iza 'project_number=' iz gornje liste. Nikad ne pišuj broj iz memorije.")
    return "\n".join(lines)


def _chat_get_employees() -> str:
    """Vraća listu zaposlenih iz baze projekata (filter: Opšte 2026, domen 9)."""
    if not PROJECTS_DB.exists():
        return "Baza projekata ne postoji. Pokreni sinhronizaciju."
    conn = sqlite3.connect(PROJECTS_DB)
    rows = conn.execute(
        "SELECT name FROM projects WHERE domain_code = '9' AND name LIKE '%2026%' ORDER BY name"
    ).fetchall()
    conn.close()
    seen: set[str] = set()
    names: list[str] = []
    for (n,) in rows:
        parts = n.split()
        if len(parts) >= 3:
            full = parts[0] + " " + parts[1]
            norm = unicodedata.normalize("NFC", full)
            if norm not in seen:
                seen.add(norm)
                names.append(full)
    names.sort()
    lines = [f"Ukupno zaposlenih: **{len(names)}**\n"]
    for i, nm in enumerate(names, 1):
        lines.append(f"{i}. {nm}")
    return "\n".join(lines)


# Podaci iz ličnih karata zaposlenih (skenirani dokumenti, február 2026)
_EMPLOYEE_ID_CARDS: list[dict] = [
    {"ime": "Aleksandar Mijailović",  "datum_rodjenja": "20.10.1982.", "jmbg": "2010982710106", "br_lk": "011188638", "adresa": "BEOGRAD, ČUKARICA, LEPOSAVE VUJOŠEVIĆ 011"},
    {"ime": "Aleksandar Živković",    "datum_rodjenja": "06.08.1985.", "jmbg": "0608985710161", "br_lk": "014839449", "adresa": "UMKA, ČUKARICA, TRINAESTOG OKTOBRA 018"},
    {"ime": "Aleksandra Gačević",     "datum_rodjenja": "20.04.1967.", "jmbg": "2004967715193", "br_lk": "010976178", "adresa": "BEOGRAD, ČUKARICA, CERSKI VENAC 023"},
    {"ime": "Ana Mitić",              "datum_rodjenja": "03.08.1986.", "jmbg": "0308986715295", "br_lk": "015063968", "adresa": "BEOGRAD, NOVI BEOGRAD, BULEVAR ARSENIJA ČARNOJEVIĆA 051/41"},
    {"ime": "Bojan Došlov",           "datum_rodjenja": "23.08.1995.", "jmbg": "2308995710031", "br_lk": "009720887", "adresa": "BEOGRAD, VOŽDOVAC, INDIRE GANDI 017"},
    {"ime": "Boško Gligić",           "datum_rodjenja": "07.01.1982.", "jmbg": "0701982710152", "br_lk": "013864602", "adresa": "BEOGRAD, ZEMUN, OHRIDSKA 001"},
    {"ime": "Boško Vuković",          "datum_rodjenja": "01.01.1990.", "jmbg": "0101990790036", "br_lk": "011532678", "adresa": "BEOGRAD, VOŽDOVAC, ŠUMADIJSKE DIVIZIJE 016/22"},
    {"ime": "Branko Stanković",       "datum_rodjenja": "01.05.1985.", "jmbg": "0105985710216", "br_lk": "008116034", "adresa": "BEOGRAD, ČUKARICA, RATKA MITROVIĆA 133"},
    {"ime": "Dalibor Gmitrović",      "datum_rodjenja": "13.02.1983.", "jmbg": "1302983760015", "br_lk": "011261404", "adresa": "BEOGRAD, ČUKARICA, LOLE RIBARA 002D/3"},
    {"ime": "Danijela Lazarević",     "datum_rodjenja": "01.08.1973.", "jmbg": "0108973715327", "br_lk": "013965691", "adresa": "BEOGRAD, VRAČAR, SAZONOVA 058/2/9"},
    {"ime": "Đorđe Ilić",            "datum_rodjenja": "17.01.1990.", "jmbg": "1701990710122", "br_lk": "012982956", "adresa": "SREMČICA, ČUKARICA, SARE BERNAR 053/POT/18"},
    {"ime": "Đorđe Virijević",       "datum_rodjenja": "12.10.1979.", "jmbg": "1210979710346", "br_lk": "014293032", "adresa": "BEOGRAD, ČUKARICA, BRAĆE MITROVIĆA 021"},
    {"ime": "Dragan Perić",           "datum_rodjenja": "29.03.1963.", "jmbg": "2903963890025", "br_lk": "012568372", "adresa": "MAČVANSKA MITROVICA, SREMSKA MITROVICA, 16. VOJVOĐANSKE DIVIZIJE 007"},
    {"ime": "Dragana Đolović",        "datum_rodjenja": "09.02.1983.", "jmbg": "0902983738544", "br_lk": "011469666", "adresa": "BEOGRAD, VOŽDOVAC, RADA NEIMARA 073/7"},
    {"ime": "Dušan Marković",         "datum_rodjenja": "02.02.1988.", "jmbg": "0202988710212", "br_lk": "012614607", "adresa": "BEOGRAD, ČUKARICA, MILORADA ĆIRIĆA 010"},
    {"ime": "Filip Milovanović",      "datum_rodjenja": "26.04.1992.", "jmbg": "2604992762047", "br_lk": "014302578", "adresa": "POŽAREVAC, POŽAREVAC, MORAVSKA 019"},
    {"ime": "Jaroslav Glagolevski",   "datum_rodjenja": "19.09.1991.", "jmbg": "ev.br.stranca 1909991060016", "br_lk": "000088430", "adresa": "BEOGRAD, ZEMUN, PRVE PRUGE 017/4/20"},
    {"ime": "Jovan Đorđević",         "datum_rodjenja": "19.08.1999.", "jmbg": "1908999781026", "br_lk": "011487621", "adresa": "RIBARI, BRUS, NEMA ULICE BB"},
    {"ime": "Jovan Ilić",             "datum_rodjenja": "31.03.1998.", "jmbg": "3103998742026", "br_lk": "009592796", "adresa": "ĆUKOVAC, VRANJE, NEMA ULICE BB"},
    {"ime": "Jovan Milisavljević",    "datum_rodjenja": "20.04.2001.", "jmbg": "2004001763817", "br_lk": "012062309", "adresa": "BISTRICA, PETROVAC NA MLAVI, SRPSKIH VLADARA 089"},
    {"ime": "Lazar Nedeljković",      "datum_rodjenja": "26.07.1997.", "jmbg": "2607997781030", "br_lk": "010448839", "adresa": "BEOGRAD, RAKOVICA, BOGDANA ŽERAJIĆA 006/19"},
    {"ime": "Luka Kraker",            "datum_rodjenja": "04.11.2001.", "jmbg": "0411001710117", "br_lk": "012087122", "adresa": "OSTRUŽNICA, ČUKARICA, DOSITEJA OBRADOVIĆA 008"},
    {"ime": "Mara Pavlović",          "datum_rodjenja": "01.03.2001.", "jmbg": "0103001765012", "br_lk": "011876857", "adresa": "SMEDEREVO, SMEDEREVO, ĐURE DANIČIĆA 107"},
    {"ime": "Marko Branković",        "datum_rodjenja": "08.12.1991.", "jmbg": "0812991710064", "br_lk": "013289641", "adresa": "ROŽANCI, BARAJEVO, DESETOG OKTOBRA 006"},
    {"ime": "Marko Burmazević",       "datum_rodjenja": "15.08.1996.", "jmbg": "1508996772028", "br_lk": "011997370", "adresa": "BEOGRAD, NOVI BEOGRAD, JURIJA GAGARINA 125/25"},
    {"ime": "Marko Đuričić",         "datum_rodjenja": "19.11.1990.", "jmbg": "1911990710208", "br_lk": "012992184", "adresa": "LISOVIĆ, BARAJEVO, RATKA JEVTIĆA 242"},
    {"ime": "Marko Gligić",           "datum_rodjenja": "06.01.1987.", "jmbg": "0601987710199", "br_lk": "010188493", "adresa": "BEOGRAD, NOVI BEOGRAD, ALEKSINAČKIH RUDARA 024/9"},
    {"ime": "Marko Ikić",             "datum_rodjenja": "17.04.2004.", "jmbg": "1704004761015", "br_lk": "011101527", "adresa": "RATARI, SMEDEREVSKA PALANKA, LAZARA STANOJEVIĆA 060"},
    {"ime": "Mićo Jarić",            "datum_rodjenja": "28.05.1966.", "jmbg": "2805966103278", "br_lk": "009959648", "adresa": "BEOGRAD, VOŽDOVAC, VOJVODE STEPE 066"},
    {"ime": "Milan Ljubojević",       "datum_rodjenja": "23.03.1965.", "jmbg": "2303965350012", "br_lk": "012991094", "adresa": "BEOGRAD, ČUKARICA, SLOBODANA PEROVIĆA 004/6"},
    {"ime": "Miloš Grbović",         "datum_rodjenja": "10.10.1996.", "jmbg": "1010996790036", "br_lk": "010494639", "adresa": "DRMANOVIĆI, NOVA VAROŠ, NEMA ULICE BB"},
    {"ime": "Miloš Jovanović",       "datum_rodjenja": "07.08.1995.", "jmbg": "0708995710032", "br_lk": "009533226", "adresa": "VRANIĆ, BARAJEVO, BRATSTVA I JEDINSTVA 258"},
    {"ime": "Miloš Pavlović",        "datum_rodjenja": "01.04.2004.", "jmbg": "0104004710258", "br_lk": "014896997", "adresa": "ROŽANCI, BARAJEVO, SELSKA 004"},
    {"ime": "Miroljub Marjanović",    "datum_rodjenja": "23.07.1952.", "jmbg": "2307952710554", "br_lk": "005568024", "adresa": "BEOGRAD, ZVEZDARA, VELJKA DUGOŠEVIĆA 006/9"},
    {"ime": "Mladen Jakovljević",     "datum_rodjenja": "29.07.1988.", "jmbg": "2907988710310", "br_lk": "009637862", "adresa": "BEOGRAD, RAKOVICA, VIDIKOVAČKI VENAC 053/47"},
    {"ime": "Momčilo Zdravković",     "datum_rodjenja": "13.06.1995.", "jmbg": "1306995710043", "br_lk": "008388941", "adresa": "BEOGRAD, SAVSKI VENAC, CARA HALIJARDA 001"},
    {"ime": "Nemanja Golubović",      "datum_rodjenja": "05.06.1994.", "jmbg": "0506994752031", "br_lk": "014685540", "adresa": "NEGOTIN, NEGOTIN, DOBRILE RADOSAVLJEVIĆ 033/4"},
    {"ime": "Nemanja Mirosavić",      "datum_rodjenja": "13.08.1991.", "jmbg": "1308991710297", "br_lk": "013863629", "adresa": "BARAJEVO, BARAJEVO, ŽIVKA STEVANOVIĆA-ŽIKICE 079B"},
    {"ime": "Nemanja Rosić",          "datum_rodjenja": "02.10.1995.", "jmbg": "0210995710026", "br_lk": "010005468", "adresa": "BEOGRAD, NOVI BEOGRAD, BULEVAR ZORANA ĐINĐIĆA 205/21"},
    {"ime": "Nikola Dragišić",        "datum_rodjenja": "20.07.2000.", "jmbg": "2007000820018", "br_lk": "012078636", "adresa": "NOVI ŽEDNIK, SUBOTICA, NIKOLE TESLE 002"},
    {"ime": "Predrag Pantić",         "datum_rodjenja": "18.10.2001.", "jmbg": "1810001710171", "br_lk": "012444483", "adresa": "BARAJEVO, BARAJEVO, DVADESETOG OKTOBRA 003"},
    {"ime": "Radomir Kerkez",         "datum_rodjenja": "13.12.1985.", "jmbg": "1312985710266", "br_lk": "012159329", "adresa": "ZUCE, VOŽDOVAC, NOVA 12 016B"},
    {"ime": "Vlada Gajić",            "datum_rodjenja": "30.04.1984.", "jmbg": "3004984710042", "br_lk": "011818283", "adresa": "RALJA, SOPOT, RADA JOVANOVIĆA 029"},
    {"ime": "Zoran Petrović",         "datum_rodjenja": "04.08.1964.", "jmbg": "0408964710189", "br_lk": "008380262", "adresa": "BAČEVAC, BARAJEVO, MILENIJE IVANOVIĆ 024"},
]


def _chat_get_id_card(ime: str) -> str:
    """Traži podatke lične karte zaposlenog po imenu ili prezimenu."""
    needle = unicodedata.normalize("NFC", ime.strip().lower())
    # normalizuj i bez dijakritika za fleksibilniju pretragu
    def _norm(s: str) -> str:
        import unicodedata as _u
        nfkd = _u.normalize("NFKD", s.lower())
        return "".join(c for c in nfkd if not _u.combining(c))
    needle_plain = _norm(ime.strip())
    results = []
    for rec in _EMPLOYEE_ID_CARDS:
        name = rec["ime"]
        if needle in unicodedata.normalize("NFC", name.lower()) or needle_plain in _norm(name):
            results.append(
                f"**{name}**\n"
                f"  Datum rođenja : {rec['datum_rodjenja']}\n"
                f"  JMBG          : {rec['jmbg']}\n"
                f"  Br. lične karte: {rec['br_lk']}\n"
                f"  Adresa        : {rec['adresa']}"
            )
    if not results:
        return f"Nema podataka lične karte za '{ime}'. Podaci postoje za 44 od 72 zaposlenih."
    return "\n\n".join(results)


def _chat_get_birthdays(n: int = 10) -> str:
    """Vraća N sledećih predstoječih rođendana zaposlenih, sortirano po broju dana."""
    from datetime import date as _date
    today = _date.today()
    entries = []
    for rec in _EMPLOYEE_ID_CARDS:
        dr = rec.get("datum_rodjenja", "").rstrip(".")
        if not dr:
            continue
        try:
            parts = dr.split(".")
            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
        except Exception:
            continue
        try:
            next_bday = _date(today.year, month, day)
        except ValueError:
            continue  # 29.02 u ne-prestupnoj godini — preskočiti
        if next_bday < today:
            try:
                next_bday = _date(today.year + 1, month, day)
            except ValueError:
                continue
        days_until = (next_bday - today).days
        age_on_bday = next_bday.year - year
        entries.append((days_until, rec["ime"], f"{day:02d}.{month:02d}.{year}.", age_on_bday))
    entries.sort(key=lambda x: x[0])
    top = entries[:n]
    lines = [f"Predstojeći rođendani (danas: {today.strftime('%d.%m.%Y.')}):\n"]
    for days, ime, dr_fmt, age in top:
        if days == 0:
            kada = "🎂 DANAS"
        elif days == 1:
            kada = "sutra"
        else:
            kada = f"za {days} dana"
        lines.append(f"- **{ime}** ({dr_fmt}) — {kada}, puni {age} god.")
    return "\n".join(lines)


def _chat_get_cars() -> str:
    """Vraća listu svih automobila sa vozačima za chat AI."""
    cars = _load_cars_from_db()
    if not cars:
        return "Nema automobila u bazi. Potrebno je pokrenuti sync."
    lines = [f"Ukupno {len(cars)} automobila:\n"]
    for c in cars:
        driver = c.get("driver", "")
        drv = f" — duži: {_cyr_to_lat(driver)}" if driver else ""
        lines.append(f"- {c['name']} ({c['unitPrice']} din/km){drv}")
    return "\n".join(lines)


def _chat_get_mileage(datum: str = "", od: str = "", do: str = "") -> str:
    """Vraća kilometraže za chat AI — za dan (datum) ili period (od/do)."""
    _ensure_mileage_table()
    conn = sqlite3.connect(PROJECTS_DB)
    conn.row_factory = sqlite3.Row

    if datum:
        try:
            d = _parse_date(datum)
        except Exception:
            return f"Neispravan datum: {datum}"
        d_str = d.strftime("%Y-%m-%d")
        rows = conn.execute(
            "SELECT car_name, start_km, end_km, amount, unit_price, total, project_name, date "
            "FROM mileage_log WHERE date = ? ORDER BY created_at",
            (d_str,),
        ).fetchall()
        conn.close()
        if not rows:
            return f"Nema kilometraže za {datum}."
        total_km = sum(r["amount"] for r in rows)
        total_din = round(sum(r["total"] for r in rows), 2)
        lines = [f"Kilometraža za {datum} ({len(rows)} unosa):\n"]
        for r in rows:
            lines.append(f"- {r['car_name']}: {r['start_km']}→{r['end_km']} = {r['amount']} km, {r['total']} din ({r['project_name'] or '?'})")
        lines.append(f"\nUkupno: {total_km} km, {total_din} din")
        return "\n".join(lines)

    # Period
    try:
        _, _, sd, ed = _period_bounds(od, do)
    except Exception:
        return "Neispravan format datuma za period."
    sd_str = sd.strftime("%Y-%m-%d")
    ed_str = ed.strftime("%Y-%m-%d")
    rows = conn.execute(
        "SELECT car_name, start_km, end_km, amount, total, project_name, date, username "
        "FROM mileage_log WHERE date >= ? AND date <= ? ORDER BY date, created_at",
        (sd_str, ed_str),
    ).fetchall()
    conn.close()
    if not rows:
        return f"Nema kilometraže za period {sd.strftime('%d.%m.%Y')} – {ed.strftime('%d.%m.%Y')}."

    total_km = sum(r["amount"] for r in rows)
    total_din = round(sum(r["total"] for r in rows), 2)
    by_day: dict[str, list] = {}
    for r in rows:
        by_day.setdefault(r["date"], []).append(r)

    lines = [f"Kilometraža {sd.strftime('%d.%m.%Y')} – {ed.strftime('%d.%m.%Y')} ({len(rows)} unosa):\n"]
    for day_str in sorted(by_day.keys()):
        d = datetime.strptime(day_str, "%Y-%m-%d").date()
        day_entries = by_day[day_str]
        day_km = sum(e["amount"] for e in day_entries)
        day_din = round(sum(e["total"] for e in day_entries), 2)
        lines.append(f"**{d.strftime('%d.%m.%Y')}** — {day_km} km, {day_din} din:")
        for e in day_entries:
            user = e["username"] if e.get("username") else ""
            lines.append(f"  - {e['car_name']}: {e['start_km']}→{e['end_km']} = {e['amount']} km ({e['project_name'] or '?'}){f' [{user}]' if user else ''}")

    lines.append(f"\n**UKUPNO: {total_km} km, {total_din} din** ({len(rows)} unosa, {len(by_day)} dana)")

    # Sumarno po automobilu
    by_car: dict[str, dict] = {}
    for r in rows:
        cn = r["car_name"]
        if cn not in by_car:
            by_car[cn] = {"km": 0, "din": 0.0, "cnt": 0}
        by_car[cn]["km"] += r["amount"]
        by_car[cn]["din"] += r["total"]
        by_car[cn]["cnt"] += 1
    lines.append("\nPo automobilu:")
    for cn in sorted(by_car.keys()):
        s = by_car[cn]
        lines.append(f"  - {cn}: {s['km']} km, {round(s['din'])} din ({s['cnt']} unosa)")

    return "\n".join(lines)


# Format taga: ⚙{"tool":"tvi_log","end_time":"18:00","comment":"..."}⚙

_CHAT_ACTION_INSTRUCTIONS = """

## Kako da izvršiš akciju

Akcioni tag koristiš ZA SVE ALATE — i za čitanje podataka (read-only) i za upis/brisanje. Dodaj ga na SAMOM KRAJU odgovora (ništa posle njega):

⚙{"tool":"IME_ALATA", "arg1":"val1", "arg2":"val2"}⚙

Za alate bez argumenata: ⚙{"tool":"tvi_birthdays"}⚙

Dostupni alati:
- tvi_log — upiši radno vreme: obavezno `end_time` (HH:MM), opcionalno `start_time`, `project_number`, `comment`, `datum`
- tvi_search — traži projekat: obavezno `pojam` (deo naziva), opcionalno `domen`
- tvi_status — zapisi za dan: opcionalno `datum` (DD.MM.YYYY)
- tvi_delete — obriši zapis: obavezno `record_id` (hex string iz konteksta — zapisi su prikazani kao `[abcdef123...] 08:00–16:00 | Projekat | komentar`; vrednost u `[...]` je tačan `record_id`)
- tvi_history — istorija perioda: opcionalno `od`, `do` (DD.MM.YYYY)
- tvi_cars — lista automobila sa vozačima (bez argumenata)
- tvi_employees — lista zaposlenih u firmi (bez argumenata); koristi kad korisnik pita "koliko ima zaposlenih", "ko radi u firmi", "lista zaposlenih"
- tvi_mileage — kilometraža: za dan `datum` (DD.MM.YYYY), ili za period `od`/`do` (DD.MM.YYYY). Koristi kad korisnik pita o pređenim km, troškovima za auto, ili kilometraži za neki period.
- tvi_licna_karta — podaci lične karte zaposlenog: `ime` (ime i/ili prezime). Vraća JMBG, broj lične karte, datum rođenja, adresu stanovanja. Podaci postoje za 44 od 72 zaposlenih.
- tvi_birthdays — predstojeći rođendani SVIH zaposlenih sortirani po datumu; opciono `n` (koliko da prikaže, podrazumevano 10). Pozivaj BEZ argumenata kad korisnik pita "kome predstoji rođendan", "ko ima sledeći rođendan", "čiji je sledeći rođendan" i slično.

PRAVILA:
1. Ako korisnik pomene projekat po imenu (ili delu naziva) — UVEK i BEZ IZUZETKA pozovi tvi_search u prvoj poruci. Ne preskačaj ovaj korak čak ni ako misliš da znaš projekat. Ne možeš znati tačan project_number bez pretrage.
2. Nakon tvi_search, backend ti vraća rezultate; tada prezentuj plan (projekat, vreme, komentar) i traži JEDNU potvrdu od korisnika — NE pozivaj tvi_log odmah
3. Tek kada korisnik da eksplicitnu potvrdu ("da", "naravno", "može"...) — pozovi tvi_log
4. `project_number` u tvi_log mora biti TAČNA cifra iza `project_number=` iz tvi_search rezultata — NIKAD ne pišuj broj koji nisi video u tvi_search rezultatima
5. U poruci korisniku posle pretrage OBAVEZNO navedi `project_number=XXXX` doslovno — taj broj mora biti vidljiv u tvom odgovoru
6. Nikad ne pozivaj tvi_log bez potvrde korisnika
13. Za brisanje: record_id su hex stringovi u `[...]` zagradi u live kontekstu ispod. Ako korisnik kaže "obriši posle 16h" — pronađi odgovarajući zapis u kontekstu, uzmi njegov `[record_id]` i pozovi tvi_delete. Ne tražи od korisnika da ti da ID — ti ga vidiš u kontekstu.
7. Tag mora biti validni JSON — bez preloma linija unutar njega
8. tvi_cars i tvi_mileage su READ-ONLY — pozivaj ih ODMAH bez potvrde kad korisnik pita o automobilima ili kilometraži
9. Kad korisnik pita "ko duži koje auto", "koji auto imam" → ODMAH pozovi tvi_cars
10. Kad korisnik pita "koliko km sam prešao", "kilometraža za dan/nedelju/mesec", "troškovi za auto" → ODMAH pozovi tvi_mileage
11. U kontekstu imaš samo SUMARNI mesečni pregled — za DETALJE po danu ili periodu UVEK koristi tvi_mileage alat
12. Za "danas" koristi tvi_mileage sa datum današnjeg datuma; za "ove nedelje" izračunaj ponedeljak i petak i koristi od/do
14. tvi_licna_karta je READ-ONLY — pozivaj ODMAH kad korisnik pita za JMBG, broj lične karte, adresu ili datum rođenja nekog zaposlenog
15. tvi_birthdays pozivaj ODMAH (bez ikakvog argumenta) kad korisnik pita ko ima sledeći/predstojeći/bliži rođendan — ovaj alat VEĆ zna sve datume za sve zaposlene, nema potrebe da tražiš ime po ime
"""


def _gemini_http_call(model: str, api_key: str, payload_dict: dict) -> dict:
    """Sinhroni Gemini REST poziv (pokreće se u asyncio.to_thread)."""
    import urllib.error
    payload = json.dumps(payload_dict).encode()
    http_req = urllib.request.Request(
        f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}",
        data=payload,
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(http_req, timeout=50) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Gemini HTTP {e.code}: {body[:800]}")


def _parse_action_tag(text: str) -> tuple[str, dict] | None:
    """Izvuče ⚙{...}⚙ tag iz teksta. Vraća (tool_name, args) ili None."""
    m = re.search(r'⚙(\{.*?\})⚙', text, re.DOTALL)
    if not m:
        return None
    try:
        data = json.loads(m.group(1))
        tool = data.pop("tool", None)
        if not tool:
            return None
        return tool, {k: v for k, v in data.items() if v not in (None, "")}
    except Exception:
        return None


@app.post("/api/chat")
async def chat(req: ChatRequest, session: dict = Depends(check_auth)):
    """Ćaskanje sa Pčelom. Akcije se izvršavaju parsiranjem tekst taga → MCP alat.

    Multi-round: ako AI pozove tvi_search kao međukorak, backend izvrši pretragu,
    ubaci rezultate i pozove AI ponovo da dovrši akciju (max 3 runde).
    """
    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(500, "GEMINI_API_KEY nije podešen u .env")
    model = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")

    try:
        user_context = await asyncio.wait_for(
            asyncio.to_thread(_build_user_context_sync, session), timeout=30.0
        )
    except Exception:
        user_context = "\n## Live podaci: trenutno nedostupni."

    system_prompt = _CHAT_SYSTEM + _CHAT_ACTION_INSTRUCTIONS + user_context

    contents = []
    for msg in req.messages[-28:]:
        gemini_role = "user" if msg.role == "user" else "model"
        contents.append({"role": gemini_role, "parts": [{"text": msg.content}]})
    if contents and contents[0]["role"] != "user":
        contents = contents[1:]

    tools_map = _get_mcp_tools_map() if _MCP_AVAILABLE else {}

    try:
        MAX_ROUNDS = 3
        for round_num in range(MAX_ROUNDS):
            resp = await asyncio.wait_for(
                asyncio.to_thread(_gemini_http_call, model, api_key, {
                    "system_instruction": {"parts": [{"text": system_prompt}]},
                    "contents": contents,
                    "generationConfig": {"temperature": 0.7, "maxOutputTokens": 2048},
                }),
                timeout=55.0,
            )

            parts = resp.get("candidates", [{}])[0].get("content", {}).get("parts", [])
            text_parts = [p["text"] for p in parts if p.get("text", "").strip()]
            ai_text = text_parts[-1].strip() if text_parts else ""

            if not ai_text:
                pf = resp.get("promptFeedback", {})
                raise RuntimeError(f"Prazan odgovor (blockReason={pf.get('blockReason','?')})")

            action = _parse_action_tag(ai_text)
            clean_text = re.sub(r'\s*⚙\{.*?\}⚙', '', ai_text, flags=re.DOTALL).rstrip()

            # ── Međukorak: tvi_search → ubaci rezultate i pozovi AI ponovo ────────
            if action and action[0] == "tvi_search" and round_num < MAX_ROUNDS - 1:
                tool_name, tool_args = action
                try:
                    search_result = await asyncio.to_thread(
                        _chat_search_projects,
                        tool_args.get("pojam", ""),
                        tool_args.get("domen", ""),
                    )
                    _log_event(
                        session["username"], "chat", "tvi_search",
                        {"args": tool_args, "round": round_num + 1}
                    )
                except Exception as e:
                    search_result = f"Greška pretrage: {e}"
                contents.append({"role": "model", "parts": [{"text": ai_text}]})
                contents.append({"role": "user", "parts": [{"text": (
                    f"[Rezultat tvi_search]\n{search_result}\n\n"
                    "Obavesti korisnika koji projekat si pronašla i šta planiraš da uradiš "
                    "(vreme, komentar). KRITIČNO: U svom odgovoru OBAVEZNO navedi tačan "
                    "project_number u formatu 'project_number=XXXX' — prepiši cifru DOSLOVNO "
                    "iz reda 'project_number=...' u rezultatima iznad. "
                    "NEMOJ sada pozvati tvi_log — sačekaj eksplicitnu potvrdu korisnika."
                )}]})
                continue

            # ── Međukorak: tvi_employees → lista zaposlenih ──────────────────────
            if action and action[0] == "tvi_employees" and round_num < MAX_ROUNDS - 1:
                try:
                    emp_result = await asyncio.to_thread(_chat_get_employees)
                    _log_event(session["username"], "chat", "tvi_employees", {"round": round_num + 1})
                except Exception as e:
                    emp_result = f"Greška: {e}"
                contents.append({"role": "model", "parts": [{"text": ai_text}]})
                contents.append({"role": "user", "parts": [{"text": (
                    f"[Rezultat tvi_employees]\n{emp_result}\n\n"
                    "Prezentuj korisniku listu zaposlenih sa ukupnim brojem."
                )}]})
                continue

            # ── Međukorak: tvi_cars → lista auta sa vozačima ─────────────────────
            if action and action[0] == "tvi_cars" and round_num < MAX_ROUNDS - 1:
                try:
                    cars_result = await asyncio.to_thread(_chat_get_cars)
                    _log_event(session["username"], "chat", "tvi_cars", {"round": round_num + 1})
                except Exception as e:
                    cars_result = f"Greška: {e}"
                contents.append({"role": "model", "parts": [{"text": ai_text}]})
                contents.append({"role": "user", "parts": [{"text": (
                    f"[Rezultat tvi_cars]\n{cars_result}\n\n"
                    "Prezentuj rezultate korisniku — lepo formatirano."
                )}]})
                continue

            # ── Međukorak: tvi_mileage → kilometraža za dan/period ───────────────
            if action and action[0] == "tvi_mileage" and round_num < MAX_ROUNDS - 1:
                tool_args = action[1]
                try:
                    mileage_result = await asyncio.to_thread(
                        _chat_get_mileage,
                        tool_args.get("datum", ""),
                        tool_args.get("od", ""),
                        tool_args.get("do", ""),
                    )
                    _log_event(session["username"], "chat", "tvi_mileage", {"args": tool_args, "round": round_num + 1})
                except Exception as e:
                    mileage_result = f"Greška: {e}"
                contents.append({"role": "model", "parts": [{"text": ai_text}]})
                contents.append({"role": "user", "parts": [{"text": (
                    f"[Rezultat tvi_mileage]\n{mileage_result}\n\n"
                    "Prezentuj rezultate korisniku — lepo formatirano, sa svim detaljima o km i troškovima."
                )}]})
                continue

            # ── Međukorak: tvi_licna_karta → podaci lične karte ──────────────────
            if action and action[0] == "tvi_licna_karta" and round_num < MAX_ROUNDS - 1:
                tool_args = action[1]
                try:
                    lk_result = await asyncio.to_thread(_chat_get_id_card, tool_args.get("ime", ""))
                    _log_event(session["username"], "chat", "tvi_licna_karta", {"args": tool_args})
                except Exception as e:
                    lk_result = f"Greška: {e}"
                contents.append({"role": "model", "parts": [{"text": ai_text}]})
                contents.append({"role": "user", "parts": [{"text": (
                    f"[Rezultat tvi_licna_karta]\n{lk_result}\n\n"
                    "Prezentuj korisniku podatke iz lične karte lepo formatirano."
                )}]})
                continue

            # ── Međukorak: tvi_birthdays → predstojeći rođendani ─────────────────
            if action and action[0] == "tvi_birthdays" and round_num < MAX_ROUNDS - 1:
                try:
                    n_bday = int(tool_args.get("n", 10)) if tool_args else 10
                    bday_result = await asyncio.to_thread(_chat_get_birthdays, n_bday)
                    _log_event(session["username"], "chat", "tvi_birthdays", {"args": tool_args})
                except Exception as e:
                    bday_result = f"Greška: {e}"
                contents.append({"role": "model", "parts": [{"text": ai_text}]})
                contents.append({"role": "user", "parts": [{"text": (
                    f"[Rezultat tvi_birthdays]\n{bday_result}\n\n"
                    "Odgovori korisniku ko ima sledeći predstojeći rođendan i kada."
                )}]})
                continue

            # ── Finalna akcija (tvi_log, tvi_delete, tvi_status, ...) ─────────────
            if action and tools_map:
                tool_name, tool_args = action
                fn = tools_map.get(tool_name)
                if fn:
                    _ctx_token = _tvi_mcp_module._session_ctx.set(session)
                    try:
                        tool_result = await fn(**tool_args)
                        result_str = str(tool_result)
                        _log_event(session["username"], "chat", tool_name, {"args": tool_args})
                        clean_text += f"\n\n✅ {result_str}"
                    except Exception as e:
                        result_str = f"Greška: {e}"
                        clean_text += f"\n\n❌ {e}"
                    finally:
                        _tvi_mcp_module._session_ctx.reset(_ctx_token)

            return {"role": "assistant", "content": clean_text or "Žao mi je, nisam uspela da odgovorim. Pokušaj ponovo."}

        # Dostignut MAX_ROUNDS bez finalne akcije
        return {"role": "assistant", "content": "Žao mi je, nisam uspela da pronađem projekat i završim akciju. Pokušaj ponovo ili navedi tačan broj projekta."}

    except asyncio.TimeoutError:
        raise HTTPException(504, "AI odgovor nije stigao na vreme. Pokušaj ponovo.")
    except Exception as e:
        raise HTTPException(500, f"Chat greška: {e}")


# ── Kilometraža / Materijalni troškovi ─────────────────────────────────────────

CARS_FILE = BASE_DIR / "cars.json"


def _load_cars_from_db() -> list[dict]:
    """Čita automobile iz SQLite. Vraća listu dicts sa ključevima kompatibilnim sa cars.json."""
    db_path = BASE_DIR / "projects" / "projects.db"
    if not db_path.exists():
        return []
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cols = [r[1] for r in conn.execute("PRAGMA table_info(cars)").fetchall()]
        has_driver = "driver" in cols
        rows = conn.execute(
            "SELECT id, code, name, type, base_unit, unit_price"
            + (", driver" if has_driver else "") +
            " FROM cars ORDER BY name"
        ).fetchall()
        conn.close()
        return [
            {
                "_id": r["id"], "code": r["code"], "name": r["name"],
                "type": r["type"], "baseUnitOfMeasure": r["base_unit"],
                "unitPrice": r["unit_price"],
                "driver": r["driver"] if has_driver else "",
            }
            for r in rows
        ]
    except Exception:
        return []


def _load_cars() -> list[dict]:
    """Prvo pokušaj SQLite, fallback na cars.json."""
    cars = _load_cars_from_db()
    if cars:
        return cars
    if CARS_FILE.exists():
        with open(CARS_FILE, encoding="utf-8") as f:
            data = json.load(f)
        return [c for c in data if c.get("unitPrice")]
    return []


def _find_km_projects(full_name: str) -> list[dict]:
    """Pronalazi projekte za kilometražu: 'Ime Prezime Automobil Kuća Posao' i 'Automobil Privatno'."""
    if not full_name or not PROJECTS_DB.exists():
        return []
    current_year = str(date.today().year)

    def _norm(t: str) -> str:
        return "".join(
            c for c in unicodedata.normalize("NFD", t.lower())
            if unicodedata.category(c) != "Mn"
        )

    fn_norm = _norm(full_name)
    fn_words = fn_norm.split()

    try:
        conn = sqlite3.connect(PROJECTS_DB)
        all_rows = conn.execute(
            "SELECT activity_number, name, domain_name, id, requests_id FROM projects "
            "WHERE name LIKE '%utomobil%' ORDER BY name"
        ).fetchall()
        conn.close()
    except Exception:
        return []

    kuca_posao = None
    privatno = None
    for r in all_rows:
        name_norm = _norm(r[1])
        if current_year not in r[1]:
            continue
        if not all(w in name_norm for w in fn_words):
            continue
        proj = {
            "activity_number": r[0] or "", "name": r[1],
            "domain": r[2], "activities_id": r[3], "requests_id": r[4],
        }
        if "privatno" in name_norm:
            privatno = proj
        elif "kuc" in name_norm or "posao" in name_norm:
            kuca_posao = proj

    result = []
    if kuca_posao:
        result.append(kuca_posao)
    if privatno:
        result.append(privatno)
    return result


@app.get("/api/cars")
async def get_cars(session: dict = Depends(check_auth)):
    cars = _load_cars()
    my_car_id = _find_my_car_id(session.get("full_name", ""), cars)

    km_projects = _find_km_projects(session.get("full_name", ""))

    return {"cars": cars, "my_car_id": my_car_id, "km_projects": km_projects}


@app.post("/api/cars/sync")
async def sync_cars(session: dict = Depends(check_auth)):
    """Ručno osvežava listu automobila iz TVI-ja."""
    try:
        result = await asyncio.to_thread(_sync_cars_blocking, session)
        return {"ok": True, **result}
    except Exception as e:
        raise HTTPException(500, f"Greška pri sync-u automobila: {e}")


@app.post("/api/timesheets/sync")
async def sync_timesheets(
    od: str = "",
    do: str = "",
    session: dict = Depends(check_auth),
):
    """Povlači radno vreme svih zaposlenih (accounts.csv) u lokalnu bazu.

    od/do: DD.MM.YYYY — default je od 01.01.2020 do danas (puni istorijat).
    Samo admin može pokrenuti.
    """
    if session["username"] != MASTER_USER:
        raise HTTPException(403, "Samo admin može pokrenuti sinhronizaciju.")

    def _run():
        return _sync_timesheets_blocking(od=od, do=do)

    try:
        result = await asyncio.to_thread(_run)
        return {"ok": True, **result}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/timesheets/sync/stream")
async def sync_timesheets_stream(
    od: str = "",
    do: str = "",
    session: dict = Depends(check_auth),
):
    """SSE stream za praćenje napretka sinhronizacije radnog vremena."""
    if session["username"] != MASTER_USER:
        raise HTTPException(403, "Samo admin.")

    q: queue.Queue = queue.Queue()

    def _cb(msg: dict):
        q.put(msg)

    def _run():
        try:
            result = _sync_timesheets_blocking(od=od, do=do, progress_cb=_cb)
            q.put({"type": "finish", **result})
        except Exception as e:
            q.put({"type": "finish", "error": str(e)})

    threading.Thread(target=_run, daemon=True).start()

    async def _gen():
        while True:
            try:
                msg = await asyncio.to_thread(q.get, True, 120)
            except Exception:
                break
            yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
            if msg.get("type") == "finish":
                break

    return StreamingResponse(
        _gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/timesheets")
async def get_timesheets(
    user_id: str = "",
    od: str = "",
    do: str = "",
    session: dict = Depends(check_auth),
):
    """Čita radno vreme iz lokalne baze. Parametri: user_id, od, do (DD.MM.YYYY)."""
    if not PROJECTS_DB.exists():
        raise HTTPException(404, "Baza ne postoji — pokreni /api/timesheets/sync.")

    conn = sqlite3.connect(PROJECTS_DB)
    try:
        # Proveri da li tabela postoji
        tbl = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='request_times'"
        ).fetchone()
        if not tbl:
            raise HTTPException(404, "Tabela request_times ne postoji — pokreni /api/timesheets/sync.")

        where, params = [], []
        if user_id:
            where.append("user_id = ?"); params.append(user_id)
        if od:
            s_ms = int(datetime.strptime(od, "%d.%m.%Y").timestamp() * 1000)
            where.append("date_ms >= ?"); params.append(s_ms)
        if do:
            e_ms = int(datetime.strptime(do, "%d.%m.%Y").replace(
                hour=23, minute=59, second=59).timestamp() * 1000)
            where.append("date_ms <= ?"); params.append(e_ms)

        sql = "SELECT id, user_id, user_name, start_ms, end_ms, date_ms, hours, price_per_hour, total, comment, project_name FROM request_times"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY date_ms, start_ms"

        rows = conn.execute(sql, params).fetchall()
        cols = ["id", "user_id", "user_name", "start_ms", "end_ms", "date_ms",
                "hours", "price_per_hour", "total", "comment", "project_name"]
        records = [dict(zip(cols, r)) for r in rows]

        employees = conn.execute(
            "SELECT user_id, full_name, price_per_hour, last_synced_at FROM employees ORDER BY full_name"
        ).fetchall() if conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='employees'"
        ).fetchone() else []

    finally:
        conn.close()

    return {
        "count": len(records),
        "records": records,
        "employees": [{"user_id": r[0], "full_name": r[1], "price_per_hour": r[2], "last_synced_at": r[3]}
                      for r in employees],
    }


def _ensure_mileage_table():
    conn = sqlite3.connect(PROJECTS_DB)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS mileage_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            car_id TEXT,
            car_name TEXT,
            start_km INTEGER,
            end_km INTEGER,
            amount INTEGER,
            unit_price REAL,
            total REAL,
            project_name TEXT,
            activities_id TEXT,
            requests_id TEXT,
            date TEXT,
            created_at TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mileage_user_date ON mileage_log(username, date)")
    conn.commit()
    conn.close()


def _save_mileage_log(username: str, car_name: str, car_id: str,
                       start_km: int, end_km: int, amount: int,
                       unit_price: float, total: float, project_name: str,
                       activities_id: str, requests_id: str, target_date: str):
    _ensure_mileage_table()
    conn = sqlite3.connect(PROJECTS_DB)
    conn.execute(
        "INSERT INTO mileage_log "
        "(username, car_id, car_name, start_km, end_km, amount, unit_price, total, "
        " project_name, activities_id, requests_id, date, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (username, car_id, car_name, start_km, end_km, amount, unit_price, total,
         project_name, activities_id, requests_id, target_date,
         datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    )
    conn.commit()
    conn.close()


def _get_mileage_log(username: str, target_date: str) -> list[dict]:
    _ensure_mileage_table()
    conn = sqlite3.connect(PROJECTS_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT id, car_name, start_km, end_km, amount, unit_price, total, "
        "project_name, created_at FROM mileage_log "
        "WHERE username = ? AND date = ? ORDER BY created_at",
        (username, target_date),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/mileage")
async def get_mileage(
    datum: str = Query(default="", description="DD.MM.YYYY, prazno = danas"),
    session: dict = Depends(check_auth),
):
    """Vraća kilometraže za dan iz lokalne baze."""
    if datum:
        try:
            d = _parse_date(datum)
        except Exception:
            raise HTTPException(400, "Neispravan format datuma.")
    else:
        d = date.today()
    date_str = d.strftime("%Y-%m-%d")
    entries = _get_mileage_log(session["username"], date_str)
    total_km = sum(e["amount"] for e in entries)
    total_din = round(sum(e["total"] for e in entries), 2)
    return {"entries": entries, "total_km": total_km, "total_din": total_din}


@app.get("/api/mileage/last-km")
async def mileage_last_km(
    car_id: str = Query(..., description="ID automobila"),
    session: dict = Depends(check_auth),
):
    """Vraća poslednju krajnju kilometražu za zadati auto i korisnika."""
    _ensure_mileage_table()
    conn = sqlite3.connect(PROJECTS_DB)
    row = conn.execute(
        "SELECT end_km FROM mileage_log WHERE username = ? AND car_id = ? "
        "ORDER BY date DESC, id DESC LIMIT 1",
        (session["username"], car_id),
    ).fetchone()
    conn.close()
    return {"last_km": row[0] if row else None}


@app.get("/api/mileage/history")
async def mileage_history(
    od: str = Query(default="", description="DD.MM.YYYY"),
    do: str = Query(default="", description="DD.MM.YYYY"),
    session: dict = Depends(check_auth),
):
    """Vraća kilometraže za period, grupisane po danu."""
    try:
        _, _, sd, ed = _period_bounds(od, do)
    except Exception:
        raise HTTPException(400, "Neispravan format datuma.")

    _ensure_mileage_table()
    sd_str = sd.strftime("%Y-%m-%d")
    ed_str = ed.strftime("%Y-%m-%d")

    conn = sqlite3.connect(PROJECTS_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT id, car_name, start_km, end_km, amount, unit_price, total, "
        "project_name, date, created_at FROM mileage_log "
        "WHERE username = ? AND date >= ? AND date <= ? ORDER BY date, created_at",
        (session["username"], sd_str, ed_str),
    ).fetchall()
    conn.close()

    by_day: dict[str, list] = {}
    for r in rows:
        d_str = r["date"]
        by_day.setdefault(d_str, []).append(dict(r))

    days = []
    total_km = 0
    total_din = 0.0
    for day_str in sorted(by_day.keys()):
        d = datetime.strptime(day_str, "%Y-%m-%d").date()
        entries = by_day[day_str]
        day_km = sum(e["amount"] for e in entries)
        day_din = round(sum(e["total"] for e in entries), 2)
        total_km += day_km
        total_din += day_din
        days.append({
            "datum": d.strftime("%d.%m.%Y"),
            "dan": WEEKDAYS[d.weekday()],
            "entries": entries,
            "total_km": day_km,
            "total_din": day_din,
        })

    return {
        "od": sd.strftime("%d.%m.%Y"),
        "do": ed.strftime("%d.%m.%Y"),
        "days": days,
        "total_km": total_km,
        "total_din": round(total_din, 2),
        "total_entries": len(rows),
    }


@app.post("/api/mileage")
async def add_mileage(req: MileageRequest, request: Request, session: dict = Depends(check_auth)):
    cars = _load_cars()
    car = next((c for c in cars if c["_id"] == req.car_id), None)
    if not car:
        raise HTTPException(400, f"Nepoznat automobil: {req.car_id}")
    if req.end_km <= req.start_km:
        raise HTTPException(400, "Krajnja kilometraža mora biti veća od početne.")

    amount = req.end_km - req.start_km
    unit_price = car["unitPrice"]

    target_date = _parse_date(req.date) if req.date else date.today()
    date_ms = int(datetime(target_date.year, target_date.month, target_date.day, 12, 0).timestamp() * 1000)

    item_obj = {
        "code": car["code"],
        "name": car["name"],
        "description": None,
        "type": car["type"],
        "baseUnitOfMeasure": car["baseUnitOfMeasure"],
        "unitPrice": unit_price,
        "unitCost": None,
        "percentage": None,
        "systemInfo": None,
        "integrationInfo": None,
        "store": None,
        "_id": {"$type": "oid", "$value": car["_id"]},
    }

    ddp = _connect_login(session)
    try:
        result = ddp.add_request_item(
            item=item_obj,
            amount=amount,
            unit_price=unit_price,
            date_ms=date_ms,
            start_km=req.start_km,
            end_km=req.end_km,
            activities_id=req.activities_id,
            requests_id=req.requests_id,
            added_by_id=session["user_id"],
        )
    finally:
        ddp.close()

    # Pronađi naziv projekta za log
    proj_name = ""
    if PROJECTS_DB.exists():
        try:
            conn = sqlite3.connect(PROJECTS_DB)
            row = conn.execute(
                "SELECT name FROM projects WHERE id = ?", (req.activities_id,)
            ).fetchone()
            conn.close()
            if row:
                proj_name = row[0]
        except Exception:
            pass

    # Sačuvaj u lokalnu bazu za pregled
    date_str = target_date.strftime("%Y-%m-%d")
    _save_mileage_log(
        username=session["username"],
        car_name=car["name"],
        car_id=req.car_id,
        start_km=req.start_km,
        end_km=req.end_km,
        amount=amount,
        unit_price=unit_price,
        total=round(amount * unit_price, 2),
        project_name=proj_name,
        activities_id=req.activities_id,
        requests_id=req.requests_id,
        target_date=date_str,
    )

    ip = _get_client_ip(request)
    _log_event(session["username"], ip, "mileage", {
        "car": car["name"],
        "start_km": req.start_km,
        "end_km": req.end_km,
        "amount": amount,
        "total": round(amount * unit_price, 2),
        "activities_id": req.activities_id,
    })

    return {
        "ok": True,
        "car": car["name"],
        "amount": amount,
        "total": round(amount * unit_price, 2),
        "unit_price": unit_price,
    }


# ── MCP u produkciji (isti alati kao tvi_mcp.py) ─────────────────────────────────

_MCP_TOOLS_META = [
    {"name": "tvi_status", "label": "Status dana", "args": [{"key": "datum", "label": "Datum (DD.MM.YYYY)", "optional": True}]},
    {"name": "tvi_status_month", "label": "Status meseca (radni dani)", "args": []},
    {"name": "tvi_log", "label": "Upis radnog vremena", "args": [
        {"key": "end_time", "label": "Do (HH:MM)", "optional": False},
        {"key": "start_time", "label": "Od (HH:MM)", "optional": True},
        {"key": "comment", "label": "Komentar", "optional": True},
        {"key": "project_number", "label": "Broj projekta", "optional": True},
        {"key": "datum", "label": "Datum (DD.MM.YYYY)", "optional": True},
    ]},
    {"name": "tvi_search", "label": "Pretraga projekata", "args": [
        {"key": "pojam", "label": "Naziv / pojam", "optional": False},
        {"key": "domen", "label": "Domen (6,8,9,10,12,13,14)", "optional": True},
    ]},
    {"name": "tvi_history", "label": "Istorija", "args": [
        {"key": "od", "label": "Od (DD.MM.YYYY)", "optional": True},
        {"key": "do", "label": "Do (DD.MM.YYYY)", "optional": True},
    ]},
    {"name": "tvi_export", "label": "Excel izveštaj", "args": [
        {"key": "od", "label": "Od (DD.MM.YYYY)", "optional": True},
        {"key": "do", "label": "Do (DD.MM.YYYY)", "optional": True},
    ]},
    {"name": "tvi_sync", "label": "Sync baze projekata", "args": []},
    {"name": "tvi_delete", "label": "Brisanje zapisa (po ID)", "args": [{"key": "record_id", "label": "ID zapisa (hex)", "optional": False}]},
    {"name": "tvi_delete_after", "label": "Brisanje svih unosa posle vremena", "args": [
        {"key": "start_time", "label": "Vreme (HH:MM) — briše unose koji počinju u ovo vreme ili posle", "optional": False},
        {"key": "datum", "label": "Datum (DD.MM.YYYY)", "optional": True},
    ]},
    {"name": "tvi_delete_day", "label": "Brisanje dana", "args": [{"key": "datum", "label": "Datum (DD.MM.YYYY)", "optional": True}]},
]

_MCP_TOOLS_MAP = None  # populated when _MCP_AVAILABLE


def _get_mcp_tools_map():
    global _MCP_TOOLS_MAP
    if not _MCP_AVAILABLE:
        return None
    if _MCP_TOOLS_MAP is None:
        _MCP_TOOLS_MAP = {
            "tvi_status": tvi_status,
            "tvi_status_month": tvi_status_month,
            "tvi_log": tvi_log,
            "tvi_search": tvi_search,
            "tvi_history": tvi_history,
            "tvi_export": tvi_export,
            "tvi_sync": tvi_sync,
            "tvi_delete": tvi_delete,
            "tvi_delete_after": tvi_delete_after,
            "tvi_delete_day": tvi_delete_day,
        }
    return _MCP_TOOLS_MAP


@app.get("/api/mcp/tools")
async def mcp_tools(session: dict = Depends(check_auth)):
    """Lista MCP alata za web UI (produkcija)."""
    if not _MCP_AVAILABLE:
        raise HTTPException(status_code=503, detail="MCP alati nisu dostupni na ovom serveru.")
    return {"tools": _MCP_TOOLS_META}


@app.post("/api/mcp/invoke")
async def mcp_invoke(req: MCPInvokeRequest, request: Request, session: dict = Depends(check_auth)):
    """Poziva MCP alat po imenu sa argumentima (produkcija)."""
    if not _MCP_AVAILABLE:
        raise HTTPException(status_code=503, detail="MCP alati nisu dostupni na ovom serveru.")
    tools_map = _get_mcp_tools_map()
    if req.tool not in tools_map:
        raise HTTPException(
            status_code=400,
            detail=f"Nepoznat alat: '{req.tool}'. Dostupni: {list(tools_map.keys())}",
        )
    fn = tools_map[req.tool]
    args = {k: v for k, v in (req.arguments or {}).items() if v != "" and v is not None}
    # Loguj sve akcione alate (ne čitanje)
    _LOGGABLE_TOOLS = {"tvi_log", "tvi_delete", "tvi_delete_after", "tvi_delete_day", "tvi_sync"}
    ip = _get_client_ip(request)
    # Postavi session kontekst da MCP funkcije koriste kredencijale ulogovanog korisnika
    _ctx_token = _tvi_mcp_module._session_ctx.set(session)
    try:
        result = await fn(**args)
        if req.tool in _LOGGABLE_TOOLS:
            _log_event(session["username"], ip, f"voice-{req.tool}", {
                "args": args,
                "result": str(result)[:300],
            })
        return {"result": result}
    except Exception as e:
        if req.tool in _LOGGABLE_TOOLS:
            _log_event(session["username"], ip, f"voice-{req.tool}", {
                "args": args,
                "error": str(e),
            }, status="error")
        return {"error": str(e), "result": None}
    finally:
        _tvi_mcp_module._session_ctx.reset(_ctx_token)


@app.get("/api/me")
async def me(session: dict = Depends(check_auth)):
    return {
        "username":  session["username"],
        "full_name": session["full_name"],
        "is_master": session["username"] == MASTER_USER,
    }


# ── Admin: logs ───────────────────────────────────────────────────────────────

@app.get("/api/admin/logs")
async def admin_logs(
    n:      int = Query(default=200, description="Broj poslednjih unosa"),
    user:   str = Query(default="",  description="Filter po korisniku"),
    action: str = Query(default="",  description="Filter po akciji (log, parse-voice, delete, auth)"),
    datum:  str = Query(default="",  description="Filter po datumu DD.MM.YYYY ili YYYY-MM-DD"),
    session: dict = Depends(check_auth),
):
    """Čita poslednjih N log unosa. Samo za master korisnika."""
    if session["username"] != MASTER_USER:
        raise HTTPException(status_code=403, detail="Pristup dozvoljen samo master korisniku.")

    if not LOG_FILE.exists():
        return {"entries": [], "total": 0}

    with _LOG_LOCK:
        lines = LOG_FILE.read_text(encoding="utf-8").splitlines()

    entries = []
    for line in reversed(lines):
        line = line.strip()
        if not line:
            continue
        try:
            e = json.loads(line)
        except Exception:
            continue
        if user   and e.get("user", "")   != user:
            continue
        if action and e.get("action", "") != action:
            continue
        if datum:
            # podrzava DD.MM.YYYY i YYYY-MM-DD
            ts = e.get("ts", "")[:10]  # "YYYY-MM-DD"
            try:
                if "." in datum:
                    d, m, y = datum.split(".")
                    filter_date = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
                else:
                    filter_date = datum
                if ts != filter_date:
                    continue
            except Exception:
                pass
        entries.append(e)
        if len(entries) >= n:
            break

    return {"entries": entries, "total": len(lines)}


@app.get("/", response_class=FileResponse)
async def webapp():
    """Servira webapp.html na root URL-u."""
    path = BASE_DIR / "webapp.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="webapp.html nije pronađen.")
    return FileResponse(path, media_type="text/html")


def _get_apk_path() -> Path | None:
    """Vraća putanju do APK fajla ako postoji."""
    candidates = [
        BASE_DIR / "Pcela.apk",
        BASE_DIR / "android-app" / "app" / "build" / "outputs" / "apk" / "release" / "app-release.apk",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


@app.get("/api/download-app")
async def download_app():
    """Preuzimanje Android aplikacije (Pčela). Servira Pcela.apk ako postoji."""
    path = _get_apk_path()
    if not path:
        raise HTTPException(
            status_code=404,
            detail="APK nije dostupan. Izgradi android-app i kopiraj app-release.apk u Pcela.apk ili u root projekta.",
        )
    return FileResponse(
        path,
        media_type="application/vnd.android.package-archive",
        filename="Pcela.apk",
    )


@app.get("/api/export")
async def export(
    od: str = Query(default="", description="DD.MM.YYYY, prazno = prvi dan meseca"),
    do: str = Query(default="", description="DD.MM.YYYY, prazno = danas"),
    session: dict = Depends(check_auth),
):
    """Generise Excel izvestaj za sve radnike iz accounts.csv i vraca kao download."""
    def _run():
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("Nedostaje openpyxl. Instaliraj: pip install openpyxl")

        if not ACCOUNTS_CSV.exists():
            raise RuntimeError(f"Fajl '{ACCOUNTS_CSV}' ne postoji.")

        accounts = []
        with open(ACCOUNTS_CSV, encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                accounts.append({
                    "username":       row["username"].strip(),
                    "password":       row["password"].strip(),
                    "user_id":        row["user_id"].strip(),
                    "full_name":      row["full_name"].strip(),
                    "price_per_hour": int(row["price_per_hour"].strip()),
                })
        if not accounts:
            raise RuntimeError("accounts.csv je prazan.")

        try:
            s_ms, e_ms, sd, ed = _period_bounds(od, do)
        except Exception:
            raise ValueError("Nevazan format datuma. Koristiti DD.MM.YYYY.")

        meteor_url = _env("METEOR_WSS_URL")
        EXPORTS_DIR.mkdir(exist_ok=True)
        filename = f"TVI_{ed.year}_{ed.month:02d}.xlsx"
        filepath = EXPORTS_DIR / filename

        h_font  = Font(bold=True, color="FFFFFF")
        h_fill  = PatternFill("solid", fgColor="1F4E79")
        t_font  = Font(bold=True)
        t_fill  = PatternFill("solid", fgColor="D9E1F2")
        r_align = Alignment(horizontal="right",  vertical="center")
        c_align = Alignment(horizontal="center", vertical="center")
        ts      = Side(style="thin")
        tborder = Border(left=ts, right=ts, top=ts, bottom=ts)

        def hdr(ws, row, n):
            for c in range(1, n + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = h_font; cell.fill = h_fill
                cell.border = tborder; cell.alignment = c_align

        def tot(ws, row, n):
            for c in range(1, n + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = t_font; cell.fill = t_fill; cell.border = tborder

        wb = openpyxl.Workbook()
        ws_p = wb.active
        ws_p.title = "Pregled"
        for ci, h in enumerate(["Radnik", "Sati", "Iznos (RSD)"], 1):
            ws_p.cell(row=1, column=ci, value=h)
        hdr(ws_p, 1, 3)

        pr = 2
        gh = 0.0
        gr = 0.0

        for acc in accounts:
            recs: list[dict] = []
            try:
                ddp = MeteorDDP(meteor_url)
                if not ddp.connect(timeout=15):
                    pass
                elif not ddp.login(acc["username"], acc["password"]):
                    ddp.close()
                else:
                    result = ddp.get_history(
                        user_id=acc["user_id"], user_name=acc["full_name"],
                        start_ms=s_ms, end_ms=e_ms
                    )
                    ddp.close()
                    if result and "result" in result and result["result"]:
                        recs = result["result"]
            except Exception:
                pass
            time.sleep(0.5)

            th = sum(r["hours"] for r in recs)
            tr = sum(r["total"] for r in recs)
            gh += th; gr += tr

            ws_p.cell(row=pr, column=1, value=acc["full_name"])
            ws_p.cell(row=pr, column=2, value=round(th, 2))
            ws_p.cell(row=pr, column=3, value=round(tr))
            for c in range(1, 4):
                ws_p.cell(row=pr, column=c).border = tborder
            ws_p.cell(row=pr, column=2).alignment = r_align
            ws_p.cell(row=pr, column=3).alignment = r_align
            pr += 1

            ws = wb.create_sheet(acc["full_name"][:31])
            dh = ["Datum", "Dan", "Projekat", "Od", "Do", "Sati", "Komentar", "Iznos (RSD)"]
            for ci, h in enumerate(dh, 1):
                ws.cell(row=1, column=ci, value=h)
            hdr(ws, 1, len(dh))
            dr = 2
            for r in sorted(recs, key=lambda x: x["startTime"]["$date"]):
                dd = datetime.fromtimestamp(r["date"]["$date"] / 1000).date()
                ws.cell(row=dr, column=1, value=f"{dd:%d.%m.%Y}")
                ws.cell(row=dr, column=2, value=WEEKDAYS[dd.weekday()])
                ws.cell(row=dr, column=3, value=r.get("requestName", ""))
                ws.cell(row=dr, column=4, value=_ms_to_str(r["startTime"]["$date"]))
                ws.cell(row=dr, column=5, value=_ms_to_str(r["endTime"]["$date"]))
                ws.cell(row=dr, column=6, value=round(r["hours"], 2))
                ws.cell(row=dr, column=7, value=r.get("comment") or "")
                ws.cell(row=dr, column=8, value=round(r["total"]))
                for c in range(1, 9):
                    ws.cell(row=dr, column=c).border = tborder
                ws.cell(row=dr, column=6).alignment = r_align
                ws.cell(row=dr, column=8).alignment = r_align
                dr += 1
            if recs:
                ws.cell(row=dr, column=5, value="UKUPNO:")
                ws.cell(row=dr, column=6, value=round(th, 2))
                ws.cell(row=dr, column=8, value=round(tr))
                tot(ws, dr, len(dh))
                ws.cell(row=dr, column=5).font = Font(bold=True)
                ws.cell(row=dr, column=6).alignment = r_align
                ws.cell(row=dr, column=8).alignment = r_align
            for ci, w in enumerate([12, 6, 45, 6, 6, 7, 38, 14], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

        ws_p.cell(row=pr, column=1, value="UKUPNO")
        ws_p.cell(row=pr, column=2, value=round(gh, 2))
        ws_p.cell(row=pr, column=3, value=round(gr))
        tot(ws_p, pr, 3)
        ws_p.cell(row=pr, column=2).alignment = r_align
        ws_p.cell(row=pr, column=3).alignment = r_align
        ws_p.column_dimensions["A"].width = 28
        ws_p.column_dimensions["B"].width = 10
        ws_p.column_dimensions["C"].width = 16

        wb.save(filepath)
        return str(filepath)

    try:
        filepath = await asyncio.to_thread(_run)
        return FileResponse(
            path=filepath,
            filename=Path(filepath).name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
