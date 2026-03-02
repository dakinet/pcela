"""
TVI Bee MCP Server — Claude direktno komunicira sa TVI platformom.

Registracija (jednom):
    claude mcp add --scope user tvi-bee -- python "C:/Users/Demo room/claude/bee/tvi_mcp.py"

Alati (tools):
    tvi_status   — upisano radno vreme za danas ili zadati datum
    tvi_log      — unos radnog vremena (sa ili bez auto-extend od zadnjeg)
    tvi_history  — istorija za period
    tvi_search   — pretraga projekata u lokalnoj bazi (bez konekcije)
    tvi_export   — Excel izvestaj za sve radnike iz accounts.csv
    tvi_sync     — osvezavanje lokalne baze projekata sa servera
"""

import asyncio
import calendar
import contextvars
import csv
import os
import sqlite3
import time
from datetime import datetime, date
from pathlib import Path

from dotenv import load_dotenv

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / ".env", override=True)

from mcp.server.fastmcp import FastMCP
from ddp_client import MeteorDDP

mcp = FastMCP(
    "tvi-bee",
    instructions=(
        "TVI platforma — time tracking za TVI (Team Inoving). "
        "Korisnik je Dalibor Gmitrović. "
        "Koristi tvi_status da vidiš šta je upisano danas pre bilo kog unosa. "
        "Format datuma je DD.MM.YYYY, format vremena je HH:MM. "
        "Ako korisnik kaže 'dopuni do X', pozovi tvi_log samo sa end_time=X (bez start_time). "
        "Za pretragu projekata koristi tvi_search — ne treba konekcija na server."
    ),
)

# ── konstante ─────────────────────────────────────────────────────────────────

WEEKDAYS    = ["Pon", "Uto", "Sre", "Cet", "Pet", "Sub", "Ned"]
PROJECTS_DB = BASE_DIR / "projects" / "projects.db"
ACCOUNTS_CSV = BASE_DIR / "accounts.csv"
EXPORTS_DIR = BASE_DIR / "exports"

# ── sync helperi (izvrsavaju se u thread-u) ───────────────────────────────────

def _env(key: str) -> str:
    val = os.getenv(key, "").strip()
    if not val:
        raise RuntimeError(f"Nedostaje '{key}' u .env fajlu.")
    return val


# Session kontekst — kada se poziva iz web app-a (api.py), ovde se upisuju
# kredencijali ulogovanog korisnika. Kada se poziva kao MCP server (Claude Code),
# ostaje None i koristi se .env fallback.
_session_ctx: contextvars.ContextVar[dict | None] = contextvars.ContextVar(
    'tvi_session', default=None
)


def _get_creds() -> tuple[str, str, str]:
    """Vraća (username, password, user_id) iz session konteksta ili .env fallback."""
    session = _session_ctx.get()
    if session:
        return session["username"], session["password"], session["user_id"]
    return _env("USERNAME"), _env("PASSWORD"), _env("USER_ID")


def _ms_to_str(ms: int) -> str:
    return datetime.fromtimestamp(ms / 1000).strftime("%H:%M")


def _time_to_ms(s: str, base: date) -> int:
    h, m = map(int, s.strip().split(":"))
    return int(datetime(base.year, base.month, base.day, h, m).timestamp() * 1000)


def _day_bounds(d: date) -> tuple[int, int]:
    start = int(datetime(d.year, d.month, d.day, 0, 0, 0).timestamp() * 1000)
    end   = int(datetime(d.year, d.month, d.day, 23, 59, 59).timestamp() * 1000)
    return start, end


def _parse_date(s: str) -> date:
    d, m, y = map(int, s.strip().split("."))
    return date(y, m, d)


def _period_bounds(od: str, do: str) -> tuple[int, int, date, date]:
    today = date.today()
    sd = _parse_date(od) if od else today.replace(day=1)
    ed = _parse_date(do) if do else today
    sms = int(datetime(sd.year, sd.month, sd.day, 0, 0, 0).timestamp() * 1000)
    ems = int(datetime(ed.year, ed.month, ed.day, 23, 59, 59).timestamp() * 1000)
    return sms, ems, sd, ed


def _oid_value(val) -> str:
    if isinstance(val, dict):
        return val.get("$value", "")
    return str(val) if val else ""


def _get_full_name(username: str) -> str:
    if ACCOUNTS_CSV.exists():
        try:
            with open(ACCOUNTS_CSV, encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    if row.get("username", "").strip() == username:
                        return row.get("full_name", "").strip()
        except Exception:
            pass
    return username


def _connect_login(username: str, password: str) -> MeteorDDP:
    ddp = MeteorDDP(_env("METEOR_WSS_URL"))
    if not ddp.connect():
        raise RuntimeError("Nije moguce povezati se na TVI server.")
    if not ddp.login(username, password):
        ddp.close()
        raise RuntimeError("Prijava nije uspela. Proveri korisnicko ime i lozinku.")
    return ddp


def _fetch_records(user_id: str, user_name: str,
                   start_ms: int, end_ms: int) -> list[dict]:
    username, password, _ = _get_creds()
    ddp = _connect_login(username, password)
    result = ddp.get_history(user_id=user_id, user_name=user_name,
                             start_ms=start_ms, end_ms=end_ms)
    ddp.close()
    if not result or "result" not in result or not result["result"]:
        return []
    return result["result"]


def _lookup_project_db(activity_number: str) -> tuple[str, str, str] | None:
    """Vraca (activities_id, requests_id, name) ili None."""
    if not PROJECTS_DB.exists():
        return None
    conn = sqlite3.connect(PROJECTS_DB)
    row = conn.execute(
        "SELECT id, requests_id, name FROM projects WHERE activity_number = ?",
        (activity_number,)
    ).fetchone()
    conn.close()
    return row or None


def _format_records(records: list[dict]) -> str:
    if not records:
        return "Nema zapisa."
    lines = []
    by_day: dict[date, list] = {}
    for r in records:
        day = datetime.fromtimestamp(r["date"]["$date"] / 1000).date()
        by_day.setdefault(day, []).append(r)

    total_h = 0.0
    total_rsd = 0.0
    for day in sorted(by_day.keys()):
        day_records = sorted(by_day[day], key=lambda x: x["startTime"]["$date"])
        dh = sum(r["hours"] for r in day_records)
        dr = sum(r["total"] for r in day_records)
        total_h += dh
        total_rsd += dr
        wd = WEEKDAYS[day.weekday()]
        lines.append(f"{wd} {day:%d.%m.%Y}  —  {dh:.2f}h  ({dr:,.0f} RSD)")
        for r in day_records:
            s = _ms_to_str(r["startTime"]["$date"])
            e = _ms_to_str(r["endTime"]["$date"])
            name = r.get("requestName", "")[:42]
            comment = (r.get("comment") or "")
            lines.append(f"  {s}-{e}  {r['hours']:.2f}h  {name}")
            if comment:
                lines.append(f"    # {comment}")
    lines.append("")
    lines.append(f"UKUPNO: {total_h:.2f}h  {total_rsd:,.0f} RSD")
    return "\n".join(lines)


# ── MCP Tools ─────────────────────────────────────────────────────────────────

@mcp.tool()
async def tvi_status(datum: str = "") -> str:
    """Prikazuje upisano radno vreme za danas (ili zadati datum).

    Args:
        datum: Datum u formatu DD.MM.YYYY. Ako je prazan, koristi se danas.
    """
    def _run():
        username, _, user_id = _get_creds()
        user_name = _get_full_name(username)
        today = date.today()
        if datum:
            try:
                today = _parse_date(datum)
            except Exception:
                raise RuntimeError(f"Nevazan format datuma: '{datum}'. Koristiti DD.MM.YYYY.")
        s_ms, e_ms = _day_bounds(today)
        records = _fetch_records(user_id, user_name, s_ms, e_ms)
        wd = WEEKDAYS[today.weekday()]
        header = f"Status za {wd} {today:%d.%m.%Y}\n{'=' * 40}"
        if not records:
            return f"{header}\nNema zapisa za ovaj dan."
        body = _format_records(records)
        records_sorted = sorted(records, key=lambda x: x["endTime"]["$date"])
        last_end = _ms_to_str(records_sorted[-1]["endTime"]["$date"])
        return f"{header}\n{body}\nPoslednji kraj: {last_end}"

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


@mcp.tool()
async def tvi_status_month() -> str:
    """Proverava da li je uneseno radno vreme za sve radne dane u tekucem mesecu.

    Vraca kratak rezime: koliko radnih dana ima mesec, za koliko je uneseno,
    i listu datuma koji nedostaju (bez unosa). Pogodno za glasovnu povratnu informaciju.
    """
    def _run():
        username, _, user_id = _get_creds()
        user_name = _get_full_name(username)
        today = date.today()
        year, month = today.year, today.month
        last_day = calendar.monthrange(year, month)[1]
        first = date(year, month, 1)
        last = date(year, month, last_day)
        s_ms = int(datetime(year, month, 1, 0, 0, 0).timestamp() * 1000)
        e_ms = int(datetime(year, month, last_day, 23, 59, 59).timestamp() * 1000)
        records = _fetch_records(user_id, user_name, s_ms, e_ms)
        days_with_records = set()
        for r in records:
            d = datetime.fromtimestamp(r["date"]["$date"] / 1000).date()
            days_with_records.add(d)
        workdays_all = []
        for day in range(1, last_day + 1):
            d = date(year, month, day)
            if d.weekday() < 5:
                workdays_all.append(d)
        # Samo radni dani do danas (uključujući danas) — budući dani se ne prikazuju
        workdays = [d for d in workdays_all if d <= today]
        covered = [d for d in workdays if d in days_with_records]
        missing = [d for d in workdays if d not in days_with_records]
        months_sr = ("", "januar", "februar", "mart", "april", "maj", "jun",
                    "jul", "avgust", "septembar", "oktobar", "novembar", "decembar")
        month_name = months_sr[month]
        out = (
            f"Status za {month_name} {year} (do {today:%d.%m.}): "
            f"uneseno za {len(covered)} od {len(workdays)} radnih dana."
        )
        if missing:
            missing_str = ", ".join(d.strftime("%d.%m.") for d in missing)
            out += f" Nedostaju: {missing_str}"
        else:
            out += " Svi radni dani su pokriveni."
        return out

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


@mcp.tool()
async def tvi_log(
    end_time: str,
    start_time: str = "",
    comment: str = "",
    project_number: str = "",
    datum: str = "",
) -> str:
    """Upisuje radno vreme na TVI platformu.

    Ako start_time nije zadat, automatski koristi kraj poslednjeg zapisa za zadati dan
    (tzv. 'dopuni do').

    Args:
        end_time: Kraj radnog vremena u formatu HH:MM (obavezno).
        start_time: Pocetak u formatu HH:MM. Ako je prazan, nastavlja od poslednjeg kraja.
        comment: Komentar za unos (npr. opis rada).
        project_number: Broj projekta (activity number, npr. '2176'). Ako je prazan,
                        koristi podrazumevani projekat iz .env (DEFAULT_ACTIVITIES_ID).
        datum: Datum u formatu DD.MM.YYYY. Ako je prazan, koristi se danas.
    """
    def _run():
        username, password, user_id = _get_creds()
        price_h   = int(_env("PRICE_PER_HOUR"))
        def_act   = _env("DEFAULT_ACTIVITIES_ID")
        def_req   = _env("DEFAULT_REQUESTS_ID")
        user_name = _get_full_name(username)
        if datum:
            try:
                today = _parse_date(datum)
            except Exception:
                raise RuntimeError(f"Nevazan format datuma: '{datum}'. Koristiti DD.MM.YYYY.")
        else:
            today = date.today()

        # Projekat
        activities_id = def_act
        requests_id   = def_req
        project_info  = f"podrazumevani ({def_act[:8]}...)"

        if project_number:
            proj = _lookup_project_db(project_number)
            if proj:
                activities_id, requests_id, proj_name = proj
                project_info = f"#{project_number} {proj_name[:40]}"
            else:
                raise RuntimeError(
                    f"Projekat sa brojem '{project_number}' nije pronađen u bazi. "
                    f"Pokreni tvi_search da nađeš tačan activityNumber."
                )

        # Vremena
        try:
            end_ms = _time_to_ms(end_time, today)
        except Exception:
            raise RuntimeError(f"Nevazan format end_time: '{end_time}'. Koristiti HH:MM.")

        if start_time:
            try:
                start_ms = _time_to_ms(start_time, today)
            except Exception:
                raise RuntimeError(f"Nevazan format start_time: '{start_time}'. Koristiti HH:MM.")
            ddp = _connect_login(username, password)
        else:
            # Auto-extend od poslednjeg kraja
            s_ms, e_ms = _day_bounds(today)
            ddp = _connect_login(username, password)
            result = ddp.get_history(user_id=user_id, user_name=user_name,
                                     start_ms=s_ms, end_ms=e_ms)
            if not result or "result" not in result or not result["result"]:
                ddp.close()
                raise RuntimeError(
                    "Nema zapisa za danas. Navedi start_time za prvi unos dana."
                )
            records = result["result"]
            start_ms = max(r["endTime"]["$date"] for r in records)

        if end_ms <= start_ms:
            ddp.close()
            raise RuntimeError(
                f"Kraj ({end_time}) mora biti posle pocetka ({_ms_to_str(start_ms)})."
            )

        hours = round((end_ms - start_ms) / 3_600_000, 4)
        total = round(hours * price_h)

        result = ddp.add_request_time(
            hours=hours,
            price_per_hour=price_h,
            comment=comment,
            engaged_user_id=user_id,
            start_ms=start_ms,
            end_ms=end_ms,
            activities_id=activities_id,
            requests_id=requests_id,
        )
        ddp.close()

        if result and "result" in result:
            record_id = _oid_value(result["result"].get("_id", {}))
            return (
                f"Uspesno upisano!\n"
                f"  Datum:    {today:%d.%m.%Y}\n"
                f"  Vreme:    {_ms_to_str(start_ms)} - {end_time}\n"
                f"  Sati:     {hours:.2f}h\n"
                f"  Iznos:    {total:,} RSD\n"
                f"  Projekat: {project_info}\n"
                f"  Komentar: {comment or '(bez komentara)'}\n"
                f"  ID:       {record_id}"
            )
        elif result and "error" in result:
            raise RuntimeError(f"Greska sa servera: {result['error']}")
        else:
            raise RuntimeError("Nije dobijen odgovor od servera.")

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


@mcp.tool()
async def tvi_delete(record_id: str) -> str:
    """Brise zapis radnog vremena po ID-u.

    Args:
        record_id: Hex ID zapisa (npr. '58a0cf98d63783d0828d1d8e').
    """
    def _run():
        username, password, _ = _get_creds()
        ddp = _connect_login(username, password)
        result = ddp.remove_request_time(record_id)
        ddp.close()
        if result is None:
            raise RuntimeError("Nije dobijen odgovor od servera.")
        if "error" in result:
            err = result["error"]
            if isinstance(err, dict) and err.get("error") == "item_already_deleted":
                raise RuntimeError(f"Zapis {record_id} ne postoji ili je vec obrisan.")
            raise RuntimeError(f"Greska sa servera: {err}")
        return f"Zapis {record_id} uspesno obrisan."

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


@mcp.tool()
async def tvi_delete_after(start_time: str, datum: str = "") -> str:
    """Brise sve zapise koji POCINJU u zadato vreme ili posle njega, za zadati dan.

    Npr. start_time=16:00 za dan 25.02.2026 obrise sve unose koji pocinju u 16:00 ili kasnije.

    Args:
        start_time: Vreme u formatu HH:MM (obavezno).
        datum: Datum u formatu DD.MM.YYYY. Ako je prazan, koristi se danas.
    """
    def _run():
        username, password, user_id = _get_creds()
        if datum:
            try:
                day = _parse_date(datum)
            except Exception:
                raise RuntimeError(f"Nevazan format datuma: '{datum}'. Koristiti DD.MM.YYYY.")
        else:
            day = date.today()
        try:
            start_ms = _time_to_ms(start_time, day)
        except Exception:
            raise RuntimeError(f"Nevazan format start_time: '{start_time}'. Koristiti HH:MM.")

        s_ms, e_ms = _day_bounds(day)
        ddp = _connect_login(username, password)
        records = ddp.get_request_time_ids_for_day(user_id, s_ms, e_ms)
        # Samo zapisi koji pocinju u start_time ili posle
        to_delete = [r for r in records if r["start_ms"] >= start_ms]
        ddp.close()

        if not to_delete:
            return f"Nema zapisa koji pocinju u {start_time} ili posle za {day:%d.%m.%Y}."

        ddp = _connect_login(username, password)
        lines = []
        for r in to_delete:
            result = ddp.remove_request_time(r["id"])
            s = _ms_to_str(r["start_ms"])
            e = _ms_to_str(r["end_ms"])
            if result and "error" not in result:
                lines.append(f"  OK  {s}-{e}  # {r.get('comment', '')[:50]}")
            else:
                err = (result or {}).get("error", {})
                lines.append(f"  GRESKA  {s}-{e}  {err}")
        ddp.close()
        return f"Obrisano {len(to_delete)} zapisa (posle {start_time}) za {day:%d.%m.%Y}:\n" + "\n".join(lines)

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


@mcp.tool()
async def tvi_delete_day(datum: str = "") -> str:
    """Brise sve zapise radnog vremena za zadati dan (ili danas).

    Args:
        datum: Datum u formatu DD.MM.YYYY. Ako je prazan, koristi se danas.
    """
    def _run():
        username, password, user_id = _get_creds()
        if datum:
            try:
                day = _parse_date(datum)
            except Exception:
                raise RuntimeError(f"Nevazan format datuma: '{datum}'. Koristiti DD.MM.YYYY.")
        else:
            day = date.today()

        s_ms, e_ms = _day_bounds(day)
        ddp = _connect_login(username, password)
        records = ddp.get_request_time_ids_for_day(user_id, s_ms, e_ms)

        if not records:
            ddp.close()
            return f"Nema zapisa za {day:%d.%m.%Y}."

        lines = []
        for r in records:
            result = ddp.remove_request_time(r["id"])
            s = _ms_to_str(r["start_ms"])
            e = _ms_to_str(r["end_ms"])
            if result and "error" not in result:
                lines.append(f"  OK  {s}-{e}  # {r['comment'][:50]}")
            else:
                err = (result or {}).get("error", {})
                lines.append(f"  GRESKA  {s}-{e}  {err}")
        ddp.close()

        return f"Brisanje za {day:%d.%m.%Y} ({len(records)} zapisa):\n" + "\n".join(lines)

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


@mcp.tool()
async def tvi_history(od: str = "", do: str = "") -> str:
    """Prikazuje istoriju radnog vremena za zadati period.

    Args:
        od: Pocetak perioda u formatu DD.MM.YYYY. Podrazumevano: prvi dan ovog meseca.
        do: Kraj perioda u formatu DD.MM.YYYY. Podrazumevano: danas.
    """
    def _run():
        username, _, user_id = _get_creds()
        user_name = _get_full_name(username)
        try:
            s_ms, e_ms, sd, ed = _period_bounds(od, do)
        except Exception:
            raise RuntimeError("Nevazan format datuma. Koristiti DD.MM.YYYY.")
        records = _fetch_records(user_id, user_name, s_ms, e_ms)
        header = f"Istorija: {sd:%d.%m.%Y} - {ed:%d.%m.%Y}\n{'=' * 40}"
        if not records:
            return f"{header}\nNema zapisa za izabrani period."
        body = _format_records(records)
        by_day = {}
        for r in records:
            day = datetime.fromtimestamp(r["date"]["$date"] / 1000).date()
            by_day.setdefault(day, []).append(r)
        return f"{header}\n{body}\nRadnih dana: {len(by_day)}"

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


@mcp.tool()
async def tvi_search(pojam: str, domen: str = "") -> str:
    """Pretrazuje projekte u lokalnoj bazi (bez konekcije na server).

    Args:
        pojam: Deo naziva projekta za pretragu (nije case-sensitive).
        domen: Opcionalni filter po kodu domena (6=Projektovanje, 8=Izvodjenje,
               9=Opste i Neradno, 10=Nadzor, 12=Servis, 13=Tehnicka Kontrola, 14=BZR).
    """
    def _run():
        if not PROJECTS_DB.exists():
            raise RuntimeError(
                f"Baza projekata ne postoji ({PROJECTS_DB}). "
                "Pokreni tvi_sync za preuzimanje projekata."
            )
        term = f"%{pojam}%"
        conn = sqlite3.connect(PROJECTS_DB)
        if domen:
            rows = conn.execute(
                "SELECT activity_number, name, domain_name, id, requests_id "
                "FROM projects WHERE name LIKE ? AND domain_code = ? "
                "ORDER BY domain_code+0, CAST(activity_number AS INTEGER)",
                (term, str(domen))
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT activity_number, name, domain_name, id, requests_id "
                "FROM projects WHERE name LIKE ? "
                "ORDER BY domain_code+0, CAST(activity_number AS INTEGER)",
                (term,)
            ).fetchall()
        conn.close()

        if not rows:
            return f"Nema rezultata za '{pojam}'."

        lines = [f"Pronadjeno {len(rows)} projekata za '{pojam}':\n"]
        cur_domain = None
        for num, name, domain_name, act_id, req_id in rows:
            if domain_name != cur_domain:
                cur_domain = domain_name
                lines.append(f"[{domain_name}]")
            lines.append(f"  #{num or '—':<6}  {name}")
        return "\n".join(lines)

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


@mcp.tool()
async def tvi_export(od: str = "", do: str = "") -> str:
    """Generise Excel izvestaj za sve radnike iz accounts.csv.

    Args:
        od: Pocetak perioda u formatu DD.MM.YYYY. Podrazumevano: prvi dan ovog meseca.
        do: Kraj perioda u formatu DD.MM.YYYY. Podrazumevano: danas.
    """
    def _run():
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("Nedostaje openpyxl. Instaliraj: pip install openpyxl")

        if not ACCOUNTS_CSV.exists():
            raise RuntimeError(f"Fajl '{ACCOUNTS_CSV}' ne postoji.")

        accounts = []
        with open(ACCOUNTS_CSV, encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                accounts.append({
                    "username":       row["username"].strip(),
                    "password":       row["password"].strip(),
                    "user_id":        row["user_id"].strip(),
                    "full_name":      row["full_name"].strip(),
                    "price_per_hour": int(row["price_per_hour"].strip()),
                })
        if not accounts:
            raise RuntimeError("accounts.csv je prazan.")

        try:
            s_ms, e_ms, sd, ed = _period_bounds(od, do)
        except Exception:
            raise RuntimeError("Nevazan format datuma. Koristiti DD.MM.YYYY.")

        meteor_url = _env("METEOR_WSS_URL")
        EXPORTS_DIR.mkdir(exist_ok=True)
        filename = f"TVI_{ed.year}_{ed.month:02d}.xlsx"
        filepath = EXPORTS_DIR / filename

        # Stilovi
        h_font   = Font(bold=True, color="FFFFFF")
        h_fill   = PatternFill("solid", fgColor="1F4E79")
        t_font   = Font(bold=True)
        t_fill   = PatternFill("solid", fgColor="D9E1F2")
        r_align  = Alignment(horizontal="right",  vertical="center")
        c_align  = Alignment(horizontal="center", vertical="center")
        ts       = Side(style="thin")
        tborder  = Border(left=ts, right=ts, top=ts, bottom=ts)

        def hdr(ws, row, n):
            for c in range(1, n + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = h_font; cell.fill = h_fill
                cell.border = tborder; cell.alignment = c_align

        def tot(ws, row, n):
            for c in range(1, n + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = t_font; cell.fill = t_fill; cell.border = tborder

        wb = openpyxl.Workbook()
        ws_p = wb.active
        ws_p.title = "Pregled"
        for ci, h in enumerate(["Radnik", "Sati", "Iznos (RSD)"], 1):
            ws_p.cell(row=1, column=ci, value=h)
        hdr(ws_p, 1, 3)

        pr = 2
        gh = 0.0
        gr = 0.0
        log_lines = []

        for acc in accounts:
            log_lines.append(f"  [{acc['full_name']}]...")
            recs: list[dict] = []
            try:
                ddp = MeteorDDP(meteor_url)
                if not ddp.connect(timeout=15):
                    log_lines.append("    GRESKA: konekcija nije uspela")
                elif not ddp.login(acc["username"], acc["password"]):
                    ddp.close()
                    log_lines.append("    GRESKA: prijava nije uspela")
                else:
                    result = ddp.get_history(
                        user_id=acc["user_id"], user_name=acc["full_name"],
                        start_ms=s_ms, end_ms=e_ms
                    )
                    ddp.close()
                    if result and "result" in result and result["result"]:
                        recs = result["result"]
                    log_lines.append(f"    {len(recs)} zapisa")
            except Exception as ex:
                log_lines.append(f"    GRESKA: {ex}")
            time.sleep(0.5)

            th = sum(r["hours"] for r in recs)
            tr = sum(r["total"] for r in recs)
            gh += th; gr += tr

            ws_p.cell(row=pr, column=1, value=acc["full_name"])
            ws_p.cell(row=pr, column=2, value=round(th, 2))
            ws_p.cell(row=pr, column=3, value=round(tr))
            for c in range(1, 4):
                ws_p.cell(row=pr, column=c).border = tborder
            ws_p.cell(row=pr, column=2).alignment = r_align
            ws_p.cell(row=pr, column=3).alignment = r_align
            pr += 1

            ws = wb.create_sheet(acc["full_name"][:31])
            dh = ["Datum", "Dan", "Projekat", "Od", "Do", "Sati", "Komentar", "Iznos (RSD)"]
            for ci, h in enumerate(dh, 1):
                ws.cell(row=1, column=ci, value=h)
            hdr(ws, 1, len(dh))
            dr = 2
            for r in sorted(recs, key=lambda x: x["startTime"]["$date"]):
                dd = datetime.fromtimestamp(r["date"]["$date"] / 1000).date()
                ws.cell(row=dr, column=1, value=f"{dd:%d.%m.%Y}")
                ws.cell(row=dr, column=2, value=WEEKDAYS[dd.weekday()])
                ws.cell(row=dr, column=3, value=r.get("requestName", ""))
                ws.cell(row=dr, column=4, value=_ms_to_str(r["startTime"]["$date"]))
                ws.cell(row=dr, column=5, value=_ms_to_str(r["endTime"]["$date"]))
                ws.cell(row=dr, column=6, value=round(r["hours"], 2))
                ws.cell(row=dr, column=7, value=r.get("comment") or "")
                ws.cell(row=dr, column=8, value=round(r["total"]))
                for c in range(1, 9):
                    ws.cell(row=dr, column=c).border = tborder
                ws.cell(row=dr, column=6).alignment = r_align
                ws.cell(row=dr, column=8).alignment = r_align
                dr += 1
            if recs:
                ws.cell(row=dr, column=5, value="UKUPNO:")
                ws.cell(row=dr, column=6, value=round(th, 2))
                ws.cell(row=dr, column=8, value=round(tr))
                tot(ws, dr, len(dh))
                ws.cell(row=dr, column=5).font = Font(bold=True)
                ws.cell(row=dr, column=6).alignment = r_align
                ws.cell(row=dr, column=8).alignment = r_align
            for ci, w in enumerate([12, 6, 45, 6, 6, 7, 38, 14], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

        ws_p.cell(row=pr, column=1, value="UKUPNO")
        ws_p.cell(row=pr, column=2, value=round(gh, 2))
        ws_p.cell(row=pr, column=3, value=round(gr))
        tot(ws_p, pr, 3)
        ws_p.cell(row=pr, column=2).alignment = r_align
        ws_p.cell(row=pr, column=3).alignment = r_align
        ws_p.column_dimensions["A"].width = 28
        ws_p.column_dimensions["B"].width = 10
        ws_p.column_dimensions["C"].width = 16

        wb.save(filepath)
        log_str = "\n".join(log_lines)
        return (
            f"Excel sacuvan: {filepath}\n\n"
            f"Period: {sd:%d.%m.%Y} - {ed:%d.%m.%Y}\n"
            f"Ukupno: {gh:.2f}h  {gr:,.0f} RSD\n\n"
            f"Detalji:\n{log_str}"
        )

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


@mcp.tool()
async def tvi_sync() -> str:
    """Osvezava lokalnu SQLite bazu projekata preuzimanjem sa TVI servera.

    Preuzima sve projekte svih domena (moze trajati nekoliko minuta).
    Potrebno pokrenuti jednom mesecno ili kad se dodaju novi projekti.
    """
    def _run():
        username, password, _ = _get_creds()

        PAGE_SIZE = 20
        PAUSE_SEC = 0.4
        OUTPUT_DIR = BASE_DIR / "projects"
        DB_PATH = OUTPUT_DIR / "projects.db"

        SYNC_DOMAINS = [
            ("6",  "Projektovanje"),
            ("8",  "IZVODJENJE"),
            ("9",  "OPSTE_I_NERADNO"),
            ("10", "NADZOR"),
            ("12", "SERVIS"),
            ("13", "TEHNICKA_KONTROLA"),
            ("14", "BZR_I_PPZ"),
        ]

        OUTPUT_DIR.mkdir(exist_ok=True)
        ddp = _connect_login(username, password)
        fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn = sqlite3.connect(DB_PATH)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS projects (
                id TEXT PRIMARY KEY, activity_number TEXT, name TEXT,
                domain_code TEXT, domain_name TEXT, requests_id TEXT, fetched_at TEXT
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_domain ON projects(domain_code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_number ON projects(activity_number)")
        conn.commit()

        totals = {}
        log_lines = []

        for dc, dn in SYNC_DOMAINS:
            all_docs = []
            page = 1
            while True:
                docs = ddp.search_activities_page(dc, "*", page=page, page_size=PAGE_SIZE)
                all_docs.extend(docs)
                if len(docs) < PAGE_SIZE:
                    break
                page += 1
                time.sleep(PAUSE_SEC)

            rows = []
            for doc in all_docs:
                req_id = _oid_value(doc.get("requests_id", ""))
                rows.append((
                    doc["_id"], str(doc.get("activityNumber", "")),
                    doc.get("name", ""), dc, dn, req_id, fetched_at,
                ))
            conn.executemany(
                "INSERT OR REPLACE INTO projects "
                "(id, activity_number, name, domain_code, domain_name, requests_id, fetched_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                rows,
            )
            conn.commit()
            totals[dn] = len(all_docs)
            log_lines.append(f"  [{dc}] {dn}: {len(all_docs)} projekata")
            time.sleep(PAUSE_SEC)

        ddp.close()
        conn.close()

        grand = sum(totals.values())
        return (
            f"Sync zavrsен: {fetched_at}\n"
            f"Baza: {DB_PATH}\n"
            f"Ukupno: {grand} projekata\n\n"
            + "\n".join(log_lines)
        )

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        return f"GRESKA: {e}"


# ── entrypoint ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    mcp.run(transport="stdio")
