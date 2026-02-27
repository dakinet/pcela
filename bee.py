"""
TVI Bee CLI — Unified CLI za TVI platformu.

Upotreba:
    python bee.py status [--datum DD.MM.YYYY]
    python bee.py log [start] [end] [--projekat BROJ] [--komentar TEKST] [--dopuni HH:MM]
    python bee.py history [--od DD.MM.YYYY] [--do DD.MM.YYYY]
    python bee.py search POJAM [--domen KOD]
    python bee.py export [--od DD.MM.YYYY] [--do DD.MM.YYYY]
    python bee.py sync
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

import argparse
import csv
import os
import sqlite3
import time
from datetime import datetime, date
from pathlib import Path

from dotenv import load_dotenv

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / ".env", override=True)

from ddp_client import MeteorDDP

# ── konstante ─────────────────────────────────────────────────────────────────

WEEKDAYS   = ["Pon", "Uto", "Sre", "Cet", "Pet", "Sub", "Ned"]
PROJECTS_DB  = BASE_DIR / "projects" / "projects.db"
ACCOUNTS_CSV = BASE_DIR / "accounts.csv"
EXPORTS_DIR  = BASE_DIR / "exports"


# ── pomocne funkcije ──────────────────────────────────────────────────────────

def env(key: str) -> str:
    val = os.getenv(key, "").strip()
    if not val:
        raise SystemExit(f"Greska: nedostaje '{key}' u .env fajlu.")
    return val


def ms_to_str(ms: int) -> str:
    return datetime.fromtimestamp(ms / 1000).strftime("%H:%M")


def time_str_to_ms(s: str, base: date) -> int:
    h, m = map(int, s.strip().split(":"))
    return int(datetime(base.year, base.month, base.day, h, m).timestamp() * 1000)


def day_bounds(d: date) -> tuple[int, int]:
    start = int(datetime(d.year, d.month, d.day, 0, 0, 0).timestamp() * 1000)
    end   = int(datetime(d.year, d.month, d.day, 23, 59, 59).timestamp() * 1000)
    return start, end


def parse_date(s: str) -> date:
    d, m, y = map(int, s.strip().split("."))
    return date(y, m, d)


def period_bounds(od: str | None, do: str | None) -> tuple[int, int, date, date]:
    today = date.today()
    start_date = parse_date(od) if od else today.replace(day=1)
    end_date   = parse_date(do) if do else today
    start_ms = int(datetime(start_date.year, start_date.month, start_date.day, 0, 0, 0).timestamp() * 1000)
    end_ms   = int(datetime(end_date.year,   end_date.month,   end_date.day,   23, 59, 59).timestamp() * 1000)
    return start_ms, end_ms, start_date, end_date


def connect_and_login(username: str, password: str) -> MeteorDDP:
    meteor_url = env("METEOR_WSS_URL")
    ddp = MeteorDDP(meteor_url)
    if not ddp.connect():
        raise SystemExit("Greska: nije moguce povezati se na server.")
    if not ddp.login(username, password):
        ddp.close()
        raise SystemExit("Greska: prijava nije uspela. Proveri korisnicko ime i lozinku.")
    return ddp


def get_records(ddp: MeteorDDP, user_id: str, user_name: str,
                start_ms: int, end_ms: int) -> list[dict]:
    result = ddp.get_history(user_id=user_id, user_name=user_name,
                             start_ms=start_ms, end_ms=end_ms)
    if not result or "result" not in result or not result["result"]:
        return []
    return result["result"]


def load_accounts() -> list[dict]:
    if not ACCOUNTS_CSV.exists():
        raise SystemExit(
            f"Greska: fajl '{ACCOUNTS_CSV}' ne postoji.\n"
            "Kreiraj ga po uzoru na accounts.csv.example ili dodaj podatke radnika."
        )
    accounts = []
    with open(ACCOUNTS_CSV, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            accounts.append({
                "username":       row["username"].strip(),
                "password":       row["password"].strip(),
                "user_id":        row["user_id"].strip(),
                "full_name":      row["full_name"].strip(),
                "price_per_hour": int(row["price_per_hour"].strip()),
            })
    if not accounts:
        raise SystemExit("Greska: accounts.csv je prazan.")
    return accounts


def get_full_name_for(username: str) -> str:
    """Vraca full_name za korisnika iz accounts.csv. Ako ne nadje, vraca username."""
    if ACCOUNTS_CSV.exists():
        try:
            with open(ACCOUNTS_CSV, encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    if row.get("username", "").strip() == username:
                        return row.get("full_name", "").strip()
        except Exception:
            pass
    return username


def lookup_project(activity_number: str) -> tuple[str, str] | None:
    """Trazi projekat po activity_number u lokalnoj bazi."""
    if not PROJECTS_DB.exists():
        return None
    conn = sqlite3.connect(PROJECTS_DB)
    row = conn.execute(
        "SELECT id, requests_id, name FROM projects WHERE activity_number = ?",
        (activity_number,)
    ).fetchone()
    conn.close()
    if row:
        print(f"  Projekat: [{activity_number}] {row[2][:55]}")
        return row[0], row[1]
    return None


def _oid_value(val) -> str:
    if isinstance(val, dict):
        return val.get("$value", "")
    return str(val) if val else ""


def display_day_records(records: list[dict]) -> None:
    """Prikazuje zapise za jedan dan, sortirane po vremenu."""
    records_sorted = sorted(records, key=lambda x: x["startTime"]["$date"])
    for r in records_sorted:
        start   = ms_to_str(r["startTime"]["$date"])
        end     = ms_to_str(r["endTime"]["$date"])
        name    = r.get("requestName", "")[:40]
        comment = (r.get("comment") or "")[:38]
        print(f"      {start}-{end}  {r['hours']:.2f}h  {name}")
        if comment:
            print(f"                   # {comment}")


# ── status ────────────────────────────────────────────────────────────────────

def cmd_status(args) -> None:
    username = env("USERNAME")
    password = env("PASSWORD")
    user_id  = env("USER_ID")
    price_h  = int(env("PRICE_PER_HOUR"))
    user_name = get_full_name_for(username)

    today = date.today()
    if args.datum:
        try:
            today = parse_date(args.datum)
        except Exception:
            raise SystemExit("Greska: nevazan format datuma. Koristiti DD.MM.YYYY.")

    start_ms, end_ms = day_bounds(today)
    weekday = WEEKDAYS[today.weekday()]

    print(f"\n  Status za {weekday} {today:%d.%m.%Y}")
    print("  " + "=" * 52)
    print("  Povezivanje...")

    ddp = connect_and_login(username, password)
    records = get_records(ddp, user_id, user_name, start_ms, end_ms)
    ddp.close()

    if not records:
        print("  Nema zapisa za ovaj dan.")
        return

    records_sorted = sorted(records, key=lambda x: x["startTime"]["$date"])
    total_hours = sum(r["hours"] for r in records)
    total_rsd   = sum(r["total"] for r in records)

    for r in records_sorted:
        start   = ms_to_str(r["startTime"]["$date"])
        end     = ms_to_str(r["endTime"]["$date"])
        name    = r.get("requestName", "")[:40]
        comment = (r.get("comment") or "")[:38]
        print(f"\n      {start} - {end}  ({r['hours']:.2f}h)")
        print(f"      {name}")
        if comment:
            print(f"      # {comment}")

    last_end_ms = max(r["endTime"]["$date"] for r in records)

    print()
    print("  " + "=" * 52)
    print(f"  Ukupno: {total_hours:.2f}h   {total_rsd:,.0f} RSD")
    print(f"  Poslednji kraj: {ms_to_str(last_end_ms)}")


# ── log ───────────────────────────────────────────────────────────────────────

def cmd_log(args) -> None:
    username   = env("USERNAME")
    password   = env("PASSWORD")
    user_id    = env("USER_ID")
    price_h    = int(env("PRICE_PER_HOUR"))
    def_act_id = env("DEFAULT_ACTIVITIES_ID")
    def_req_id = env("DEFAULT_REQUESTS_ID")
    user_name  = get_full_name_for(username)

    today = date.today()
    activities_id = def_act_id
    requests_id   = def_req_id

    # Projekat po broju aktivnosti
    if args.projekat:
        proj = lookup_project(str(args.projekat))
        if proj:
            activities_id, requests_id = proj
        else:
            print(f"  Upozorenje: projekat #{args.projekat} nije nadjen u bazi. Koristim podrazumevani.")

    # Odredjivanje vremenskog intervala
    if args.dopuni:
        # Preuzmi poslednji kraj za danas, pa dopuni do zadatog vremena
        print("  Preuzimanje poslednjeg zapisa za danas...")
        ddp = connect_and_login(username, password)
        s_ms, e_ms = day_bounds(today)
        records = get_records(ddp, user_id, user_name, s_ms, e_ms)

        if not records:
            ddp.close()
            raise SystemExit("  Nema zapisa za danas. Koristiti 'log START END' za prvi unos.")

        last_end_ms = max(r["endTime"]["$date"] for r in records)
        start_ms_log = last_end_ms

        try:
            end_ms_log = time_str_to_ms(args.dopuni, today)
        except Exception:
            ddp.close()
            raise SystemExit("Greska: nevazan format vremena. Koristiti HH:MM.")

        if end_ms_log <= start_ms_log:
            ddp.close()
            raise SystemExit(
                f"Greska: kraj ({args.dopuni}) nije posle poslednjeg kraja ({ms_to_str(last_end_ms)})."
            )

        print(f"  Dopunjava od {ms_to_str(start_ms_log)} do {args.dopuni}")

    elif args.start and args.end:
        ddp = None
        try:
            start_ms_log = time_str_to_ms(args.start, today)
            end_ms_log   = time_str_to_ms(args.end, today)
        except Exception:
            raise SystemExit("Greska: nevazan format vremena. Koristiti HH:MM.")

    else:
        # Interaktivni mod
        ddp = None
        print("  Interaktivni unos radnog vremena\n")
        s_str = input("  Pocetak (HH:MM): ").strip()
        e_str = input("  Kraj    (HH:MM): ").strip()
        try:
            start_ms_log = time_str_to_ms(s_str, today)
            end_ms_log   = time_str_to_ms(e_str, today)
        except Exception:
            raise SystemExit("Greska: nevazan format vremena.")

    if "ddp" not in dir() or ddp is None:
        ddp = None  # ce se kreirati pri slanju

    hours = round((end_ms_log - start_ms_log) / 3_600_000, 4)
    if hours <= 0:
        if ddp:
            ddp.close()
        raise SystemExit("Greska: kraj mora biti posle pocetka.")

    total = round(hours * price_h)

    comment = args.komentar or ""
    if not comment:
        comment = input("  Komentar: ").strip()

    print(f"\n  Spreman unos:")
    print(f"  {'─' * 42}")
    print(f"  Datum:    {today:%d.%m.%Y}")
    print(f"  Vreme:    {ms_to_str(start_ms_log)} - {ms_to_str(end_ms_log)}")
    print(f"  Sati:     {hours:.2f}h")
    print(f"  Iznos:    {total:,} RSD")
    if comment:
        print(f"  Komentar: {comment[:50]}")
    print(f"  {'─' * 42}")

    if input("\n  Poslati? (da/ne): ").strip().lower() != "da":
        print("  Otkazano.")
        if ddp:
            ddp.close()
        return

    if ddp is None:
        print("  Povezivanje...")
        ddp = connect_and_login(username, password)

    print("  Slanje...")
    result = ddp.add_request_time(
        hours=hours,
        price_per_hour=price_h,
        comment=comment,
        engaged_user_id=user_id,
        start_ms=start_ms_log,
        end_ms=end_ms_log,
        activities_id=activities_id,
        requests_id=requests_id,
    )
    ddp.close()

    if result and "result" in result:
        record_id = _oid_value(result["result"].get("_id", {}))
        print(f"\n  Uspesno! Upisano {hours:.2f}h za {today:%d.%m.%Y}.")
        if record_id:
            print(f"  ID zapisa: {record_id}")
    elif result and "error" in result:
        print(f"\n  Greska sa servera: {result['error']}")
    else:
        print("\n  Greska: nije dobijen odgovor od servera.")


# ── history ───────────────────────────────────────────────────────────────────

def cmd_history(args) -> None:
    username  = env("USERNAME")
    password  = env("PASSWORD")
    user_id   = env("USER_ID")
    price_h   = int(env("PRICE_PER_HOUR"))
    user_name = get_full_name_for(username)

    try:
        start_ms, end_ms, start_date, end_date = period_bounds(args.od, args.do)
    except Exception:
        raise SystemExit("Greska: nevazan format datuma. Koristiti DD.MM.YYYY.")

    print(f"\n  Istorija: {start_date:%d.%m.%Y} - {end_date:%d.%m.%Y}")
    print("  Povezivanje...")

    ddp = connect_and_login(username, password)
    records = get_records(ddp, user_id, user_name, start_ms, end_ms)
    ddp.close()

    if not records:
        print("  Nema zapisa za izabrani period.")
        return

    by_day: dict[date, list] = {}
    for r in records:
        day = datetime.fromtimestamp(r["date"]["$date"] / 1000).date()
        by_day.setdefault(day, []).append(r)

    total_hours = 0.0
    total_rsd   = 0.0

    print()
    for day in sorted(by_day.keys()):
        day_records = by_day[day]
        day_hours = sum(r["hours"] for r in day_records)
        day_total = sum(r["total"] for r in day_records)
        total_hours += day_hours
        total_rsd   += day_total

        weekday = WEEKDAYS[day.weekday()]
        print(f"  {weekday} {day:%d.%m.%Y}  —  {day_hours:.2f}h  ({day_total:,.0f} RSD)")
        display_day_records(day_records)

    print()
    print("  " + "=" * 52)
    print(f"  UKUPNO:  {total_hours:.2f}h   {total_rsd:,.0f} RSD")
    print(f"  Radnih dana: {len(by_day)}")
    print("  " + "=" * 52)


# ── search ────────────────────────────────────────────────────────────────────

def cmd_search(args) -> None:
    if not PROJECTS_DB.exists():
        raise SystemExit(
            f"Greska: baza projekata ne postoji ({PROJECTS_DB}).\n"
            "Pokreni 'python bee.py sync' za preuzimanje projekata."
        )

    term = f"%{args.pojam}%"
    conn = sqlite3.connect(PROJECTS_DB)

    if args.domen:
        rows = conn.execute(
            "SELECT activity_number, name, domain_name, id, requests_id "
            "FROM projects WHERE name LIKE ? AND domain_code = ? "
            "ORDER BY domain_code+0, CAST(activity_number AS INTEGER)",
            (term, str(args.domen))
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT activity_number, name, domain_name, id, requests_id "
            "FROM projects WHERE name LIKE ? "
            "ORDER BY domain_code+0, CAST(activity_number AS INTEGER)",
            (term,)
        ).fetchall()
    conn.close()

    if not rows:
        print(f"\n  Nema rezultata za '{args.pojam}'.")
        return

    print(f"\n  Pronadjeno {len(rows)} projekata za '{args.pojam}':\n")
    cur_domain = None
    for num, name, domain_name, act_id, req_id in rows:
        if domain_name != cur_domain:
            cur_domain = domain_name
            print(f"  [{domain_name}]")
        num_str = str(num) if num else "—"
        print(f"    #{num_str:<6}  {name[:58]}")
    print()


# ── export ────────────────────────────────────────────────────────────────────

def cmd_export(args) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("Greska: nedostaje openpyxl. Instaliraj: pip install openpyxl")

    accounts = load_accounts()

    try:
        start_ms, end_ms, start_date, end_date = period_bounds(args.od, args.do)
    except Exception:
        raise SystemExit("Greska: nevazan format datuma. Koristiti DD.MM.YYYY.")

    meteor_url = env("METEOR_WSS_URL")
    EXPORTS_DIR.mkdir(exist_ok=True)

    filename = f"TVI_{end_date.year}_{end_date.month:02d}.xlsx"
    filepath = EXPORTS_DIR / filename

    print(f"\n  Export: {start_date:%d.%m.%Y} - {end_date:%d.%m.%Y}")
    print(f"  Radnika: {len(accounts)}")
    print(f"  Fajl:    {filepath}\n")

    # Stilovi
    header_font      = Font(bold=True, color="FFFFFF")
    header_fill      = PatternFill("solid", fgColor="1F4E79")
    total_font       = Font(bold=True)
    total_fill       = PatternFill("solid", fgColor="D9E1F2")
    right_align      = Alignment(horizontal="right",  vertical="center")
    center_align     = Alignment(horizontal="center", vertical="center")
    thin_side        = Side(style="thin")
    thin_border      = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def apply_header(ws, row, num_cols):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.border    = thin_border
            cell.alignment = center_align

    def apply_total(ws, row, num_cols):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font   = total_font
            cell.fill   = total_fill
            cell.border = thin_border

    wb = openpyxl.Workbook()
    ws_pregled = wb.active
    ws_pregled.title = "Pregled"

    # Pregled — zaglavlje
    pregled_headers = ["Radnik", "Sati", "Iznos (RSD)"]
    for ci, h in enumerate(pregled_headers, 1):
        ws_pregled.cell(row=1, column=ci, value=h)
    apply_header(ws_pregled, 1, len(pregled_headers))

    pregled_row = 2
    grand_hours = 0.0
    grand_rsd   = 0.0

    # Preuzimanje podataka za svakog radnika
    all_data: list[tuple[dict, list[dict]]] = []

    for acc in accounts:
        print(f"  [{acc['full_name']}] Povezivanje...", end="", flush=True)
        records: list[dict] = []
        try:
            ddp = MeteorDDP(meteor_url)
            if not ddp.connect(timeout=15):
                print(" GRESKA: nije moguce povezati se. Preskacam.")
            elif not ddp.login(acc["username"], acc["password"]):
                ddp.close()
                print(" GRESKA: prijava nije uspela. Preskacam.")
            else:
                records = get_records(ddp, acc["user_id"], acc["full_name"], start_ms, end_ms)
                ddp.close()
                h_sum = sum(r["hours"] for r in records)
                print(f" {len(records)} zapisa ({h_sum:.2f}h)")
        except Exception as e:
            print(f" GRESKA: {e}. Preskacam.")
        all_data.append((acc, records))
        time.sleep(0.5)

    # Popunjavanje Excel fajla
    det_headers = ["Datum", "Dan", "Projekat", "Od", "Do", "Sati", "Komentar", "Iznos (RSD)"]

    for acc, records in all_data:
        total_hours = sum(r["hours"] for r in records)
        total_rsd   = sum(r["total"] for r in records)
        grand_hours += total_hours
        grand_rsd   += total_rsd

        # Pregled red
        ws_pregled.cell(row=pregled_row, column=1, value=acc["full_name"])
        ws_pregled.cell(row=pregled_row, column=2, value=round(total_hours, 2))
        ws_pregled.cell(row=pregled_row, column=3, value=round(total_rsd))
        for c in range(1, 4):
            ws_pregled.cell(row=pregled_row, column=c).border = thin_border
        ws_pregled.cell(row=pregled_row, column=2).alignment = right_align
        ws_pregled.cell(row=pregled_row, column=3).alignment = right_align
        pregled_row += 1

        # Sheet za radnika
        safe_name = acc["full_name"][:31]
        ws = wb.create_sheet(safe_name)

        for ci, h in enumerate(det_headers, 1):
            ws.cell(row=1, column=ci, value=h)
        apply_header(ws, 1, len(det_headers))

        det_row = 2
        records_sorted = sorted(records, key=lambda x: x["startTime"]["$date"])

        for r in records_sorted:
            day_dt  = datetime.fromtimestamp(r["date"]["$date"] / 1000).date()
            start_s = ms_to_str(r["startTime"]["$date"])
            end_s   = ms_to_str(r["endTime"]["$date"])
            name    = r.get("requestName", "")
            comment = r.get("comment") or ""
            hours   = r["hours"]
            iznos   = r["total"]
            weekday = WEEKDAYS[day_dt.weekday()]

            ws.cell(row=det_row, column=1, value=f"{day_dt:%d.%m.%Y}")
            ws.cell(row=det_row, column=2, value=weekday)
            ws.cell(row=det_row, column=3, value=name)
            ws.cell(row=det_row, column=4, value=start_s)
            ws.cell(row=det_row, column=5, value=end_s)
            ws.cell(row=det_row, column=6, value=round(hours, 2))
            ws.cell(row=det_row, column=7, value=comment)
            ws.cell(row=det_row, column=8, value=round(iznos))

            for c in range(1, 9):
                ws.cell(row=det_row, column=c).border = thin_border
            ws.cell(row=det_row, column=6).alignment = right_align
            ws.cell(row=det_row, column=8).alignment = right_align
            det_row += 1

        # Footer sa ukupno
        if records:
            ws.cell(row=det_row, column=5, value="UKUPNO:")
            ws.cell(row=det_row, column=6, value=round(total_hours, 2))
            ws.cell(row=det_row, column=8, value=round(total_rsd))
            apply_total(ws, det_row, len(det_headers))
            ws.cell(row=det_row, column=5).font = Font(bold=True)
            ws.cell(row=det_row, column=6).alignment = right_align
            ws.cell(row=det_row, column=8).alignment = right_align

        # Sirine kolona: Datum, Dan, Projekat, Od, Do, Sati, Komentar, Iznos
        col_widths = [12, 6, 45, 6, 6, 7, 40, 14]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # Pregled — UKUPNO red
    ws_pregled.cell(row=pregled_row, column=1, value="UKUPNO")
    ws_pregled.cell(row=pregled_row, column=2, value=round(grand_hours, 2))
    ws_pregled.cell(row=pregled_row, column=3, value=round(grand_rsd))
    apply_total(ws_pregled, pregled_row, 3)
    ws_pregled.cell(row=pregled_row, column=2).alignment = right_align
    ws_pregled.cell(row=pregled_row, column=3).alignment = right_align

    # Sirine kolona za Pregled
    ws_pregled.column_dimensions["A"].width = 28
    ws_pregled.column_dimensions["B"].width = 10
    ws_pregled.column_dimensions["C"].width = 16

    wb.save(filepath)
    print(f"\n  Sacuvano: {filepath}")
    print(f"  Ukupno: {grand_hours:.2f}h   {grand_rsd:,.0f} RSD")


# ── sync ──────────────────────────────────────────────────────────────────────

def cmd_sync(args) -> None:
    username   = env("USERNAME")
    password   = env("PASSWORD")

    PAGE_SIZE  = 20
    PAUSE_SEC  = 0.4
    OUTPUT_DIR = BASE_DIR / "projects"
    DB_PATH    = OUTPUT_DIR / "projects.db"

    SYNC_DOMAINS = [
        ("6",  "Projektovanje"),
        ("8",  "IZVODJENJE"),
        ("9",  "OPSTE_I_NERADNO"),
        ("10", "NADZOR"),
        ("12", "SERVIS"),
        ("13", "TEHNICKA_KONTROLA"),
        ("14", "BZR_I_PPZ"),
    ]

    OUTPUT_DIR.mkdir(exist_ok=True)

    print("\n  TVI — Osvezavanje baze projekata")
    print("  " + "=" * 44)
    print("  Povezivanje...")

    ddp = connect_and_login(username, password)
    print("  OK\n")

    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id              TEXT PRIMARY KEY,
            activity_number TEXT,
            name            TEXT,
            domain_code     TEXT,
            domain_name     TEXT,
            requests_id     TEXT,
            fetched_at      TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_domain ON projects(domain_code)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_number ON projects(activity_number)")
    conn.commit()

    totals: dict[str, int] = {}

    for domain_code, domain_name in SYNC_DOMAINS:
        print(f"  [{domain_code:>2}] {domain_name}")
        all_docs: list[dict] = []
        page = 1

        while True:
            print(f"    str. {page:3d} ... ", end="", flush=True)
            docs = ddp.search_activities_page(domain_code, "*", page=page, page_size=PAGE_SIZE)
            print(f"{len(docs):3d} projekata")
            all_docs.extend(docs)
            if len(docs) < PAGE_SIZE:
                break
            page += 1
            time.sleep(PAUSE_SEC)

        rows = []
        for doc in all_docs:
            req_id = _oid_value(doc.get("requests_id", ""))
            rows.append((
                doc["_id"],
                str(doc.get("activityNumber", "")),
                doc.get("name", ""),
                domain_code,
                domain_name,
                req_id,
                fetched_at,
            ))
        conn.executemany(
            "INSERT OR REPLACE INTO projects "
            "(id, activity_number, name, domain_code, domain_name, requests_id, fetched_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            rows,
        )
        conn.commit()
        totals[domain_name] = len(all_docs)
        print(f"  → {len(all_docs)} projekata\n")
        time.sleep(PAUSE_SEC)

    ddp.close()
    conn.close()

    grand = sum(totals.values())
    print(f"  Gotovo! Preuzeto {grand} projekata ukupno.")
    print(f"  Baza: {DB_PATH}")


# ── argparse i main ───────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="bee",
        description="TVI Bee CLI — Unified CLI za TVI platformu",
    )
    sub = parser.add_subparsers(dest="cmd", metavar="<komanda>")
    sub.required = True

    # status
    p = sub.add_parser("status", help="Prikaz zapisa za danas (ili zadati datum)")
    p.add_argument("--datum", metavar="DD.MM.YYYY", help="Datum (podrazumevano: danas)")

    # log
    p = sub.add_parser("log", help="Unos radnog vremena")
    p.add_argument("start", nargs="?", metavar="HH:MM", help="Pocetak")
    p.add_argument("end",   nargs="?", metavar="HH:MM", help="Kraj")
    p.add_argument("--dopuni",   metavar="HH:MM",  help="Dopuni od poslednjeg zapisa do zadatog vremena")
    p.add_argument("--projekat", metavar="BROJ",   help="Broj projekta (activity number, npr. 2176)")
    p.add_argument("--komentar", metavar="TEKST",  help="Komentar")

    # history
    p = sub.add_parser("history", help="Istorija radnog vremena")
    p.add_argument("--od", metavar="DD.MM.YYYY", help="Pocetak perioda (podrazumevano: 1. ovog meseca)")
    p.add_argument("--do", metavar="DD.MM.YYYY", help="Kraj perioda (podrazumevano: danas)")

    # search
    p = sub.add_parser("search", help="Pretraga projekata u lokalnoj bazi")
    p.add_argument("pojam", help="Pojam za pretragu (deo naziva projekta)")
    p.add_argument("--domen", metavar="KOD", help="Filter po kodu domena (6/8/9/10/12/13/14)")

    # export
    p = sub.add_parser("export", help="Generisanje Excel izvestaja za sve radnike iz accounts.csv")
    p.add_argument("--od", metavar="DD.MM.YYYY", help="Pocetak perioda")
    p.add_argument("--do", metavar="DD.MM.YYYY", help="Kraj perioda")

    # sync
    sub.add_parser("sync", help="Osvezavanje lokalne baze projekata sa servera")

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    cmd_map = {
        "status":  cmd_status,
        "log":     cmd_log,
        "history": cmd_history,
        "search":  cmd_search,
        "export":  cmd_export,
        "sync":    cmd_sync,
    }
    cmd_map[args.cmd](args)


if __name__ == "__main__":
    main()
